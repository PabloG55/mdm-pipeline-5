"""
MDM Full Pipeline — v3 Unified Parallel
========================================
STEP 1  Address pre-processing  (clean, translate, normalize)
STEP 2  Address validation      (Azure Maps address search + correction fallbacks)
STEP 3  Filter                  (score=1.0 + VALID  →  company match)
STEP 4  Company match           (Tier1 Fuzzy → Tier2 POI → Tier3 OpenAI web search)

All 4 steps run per-record inside ONE worker node.
10 worker threads process records in parallel.
Records are independent — each thread owns one record end-to-end.

Usage
-----
  pip install httpx pandas tqdm python-dotenv openai rapidfuzz langdetect deep-translator openpyxl
  python mdm_pipeline_v9.py                         # full run
  python mdm_pipeline_v9.py --test                  # first 10 records only (worker assignment test)
  python mdm_pipeline_v9.py --input my_data.csv     # custom input
"""

import os, re, json, time, logging, csv, argparse, unicodedata
from typing import Optional, Dict, List
from concurrent.futures import ThreadPoolExecutor, as_completed
from queue import Queue, Empty
import threading

import httpx
import pandas as pd
from tqdm import tqdm
from dotenv import load_dotenv
from openai import OpenAI

try:
    from rapidfuzz import fuzz as rfuzz
    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False

try:
    from langdetect import detect, LangDetectException
    HAVE_LANGDETECT = True
except ImportError:
    HAVE_LANGDETECT = False

try:
    from deep_translator import GoogleTranslator
    HAVE_TRANSLATOR = True
except ImportError:
    HAVE_TRANSLATOR = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# ═══════════════════════════════════════════════════════════════════════════════
# ██  CONFIG  — edit here
# ═══════════════════════════════════════════════════════════════════════════════
AZURE_MAPS_KEY = os.getenv("AZURE_MAPS_KEY", "").strip()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()

INPUT_CSV       = "100_sample_MDM.csv"      # ← your raw input
OUTPUT_CSV      = "mdm_pipeline_results.csv"

MAX_WORKERS     = 20     # parallel worker threads
ADDR_DELAY      = 0.0    # sleep between address API calls (set 0.2 if hitting rate limits)
COMPANY_DELAY   = 0.0    # sleep between company API calls
BATCH_SIZE      = 10     # print summary banner every N completed records
TEST_MODE_ROWS  = 10     # how many rows to run in --test mode

# Address validation
DELIMITER            = "|#|#"
ADDR_FIELDS          = ["street", "city", "state", "country", "zip"]
ADDR_CONFIDENCE_THR  = 0.6
ADDRESS_URL          = "https://atlas.microsoft.com/search/address/json"

# Company match
FUZZY_URL            = "https://atlas.microsoft.com/search/fuzzy/json"
POI_URL              = "https://atlas.microsoft.com/search/poi/json"
GEOCODE_URL          = "https://atlas.microsoft.com/search/address/json"
MODEL_NAME           = "gpt-4.1-mini"
NAME_SIM_THRESHOLD   = 0.72
ADDR_SIM_THRESHOLD   = 0.55
REGION_THRESHOLDS    = {
    "CN":0.60,"JP":0.60,"KR":0.60,"SA":0.62,"AE":0.62,
    "TR":0.65,"RU":0.65,"IN":0.65,
}

# ═══════════════════════════════════════════════════════════════════════════════
# ██  OUTPUT COLUMNS
# ═══════════════════════════════════════════════════════════════════════════════
OUTPUT_COLUMNS = [
    # identifiers
    "MDM_KEY","SOURCE_NAME","FULL_ADDRESS_RAW",
    # address pre-processing
    "street","street_original","city","state","country","zip",
    "address_for_api","detected_lang","has_non_latin","street_translated",
    "mojibake_rescued","mojibake_fields",
    "pre_flags","skip_api",
    # address validation
    "addr_final_status","addr_correction_strategy",
    "val_score","val_match_type",
    "val_returned_street","val_returned_city","val_returned_state",
    "val_returned_country","val_returned_zip","val_returned_freeform",
    "val_lat","val_lon",
    # filter decision
    "routing_decision",
    "sent_to_company_match",
    # company match
    "match_status","company_exists","tier_used",
    "canonical_name","website",
    "all_known_locations","locations_in_state",
    "best_address","best_street","best_city","best_state","best_zip",
    "mdm_address_occupant","nearest_location_reasoning",
    "address_match","ai_confidence","ai_evidence",
    "fuzzy_match_name","fuzzy_similarity","fuzzy_found_address",
    "poi_match_name","poi_similarity",
    # per-record confidence (every record, every tier — explains the score)
    "confidence_score","confidence_reason",
    # enrichment (firmographics — populated when company is verified)
    "legal_name","parent_company","domestic_ultimate","global_ultimate",
    "is_headquarters","email_domain",
    "naics_code","naics_description","sic_code","sic_description",
    "employee_count_range","revenue_range","year_established","industry",
    "enrich_source",
    # sources — URLs the AI grounded against (from web_search_preview citations)
    "sources","sources_verify","sources_enrich","sources_hq",
    # headquarters resolution (climbs parent/ultimate hierarchy + website fallback)
    "headquarters_address","headquarters_source",
    # meta
    "worker_id","pipeline_latency_sec",
]

if not OPENAI_API_KEY: raise ValueError("OPENAI_API_KEY missing from .env")
if not AZURE_MAPS_KEY: raise ValueError("AZURE_MAPS_KEY missing from .env")
oai_client = OpenAI(api_key=OPENAI_API_KEY)

# ═══════════════════════════════════════════════════════════════════════════════
# ██  THREAD-LOCAL HTTP CLIENT  (connection pooling — 2-3x faster Azure calls)
# ═══════════════════════════════════════════════════════════════════════════════
_tl = threading.local()

def http() -> httpx.Client:
    if not hasattr(_tl, "client"):
        _tl.client = httpx.Client(timeout=15, http2=False)
    return _tl.client

def safe_get(url: str, params: dict, retries: int = 3) -> httpx.Response:
    for attempt in range(retries):
        try:
            r = http().get(url, params=params)
            r.raise_for_status()
            return r
        except Exception as e:
            if attempt == retries - 1: raise
            time.sleep(0.5 * (2 ** attempt))

# ═══════════════════════════════════════════════════════════════════════════════
# ██  CACHE  (skip repeat company+address pairs)
# ═══════════════════════════════════════════════════════════════════════════════
_cache: Dict[tuple, Dict] = {}
_cache_lock = threading.Lock()

def cache_get(k):
    with _cache_lock: return _cache.get(k)

def cache_set(k, v):
    with _cache_lock: _cache[k] = v

# Name-first cache — keyed by (cleaned_name.lower(), country.upper())
# High hit rate expected: many MDM records share companies (e.g., Tj Maxx 01607, 01608)
_name_cache: Dict[tuple, Dict] = {}
_name_cache_lock = threading.Lock()

def name_cache_get(k):
    with _name_cache_lock: return _name_cache.get(k)

def name_cache_set(k, v):
    with _name_cache_lock: _name_cache[k] = v

# Enrichment cache — firmographic fields are company-level, not address-specific.
# Keyed by (canonical_name.lower(), country.upper()). One OpenAI call per unique
# company in the run; subsequent records with same name+country reuse it.
# Populated by both _openai_enrich (Tier 1/2 path) and _openai_verify (Tier 3 path).
_enrich_cache: Dict[tuple, Dict] = {}
_enrich_cache_lock = threading.Lock()

def enrich_cache_get(k):
    with _enrich_cache_lock: return _enrich_cache.get(k)

def enrich_cache_set(k, v):
    with _enrich_cache_lock: _enrich_cache[k] = v

# Headquarters cache — keyed by (target_name.lower(), country.upper()) where target
# is the parent / domestic_ultimate / global_ultimate / canonical_name we searched.
# Many MDM records share parents → high hit rate on same-parent groups.
_hq_cache: Dict[tuple, Dict] = {}
_hq_cache_lock = threading.Lock()

def hq_cache_get(k):
    with _hq_cache_lock: return _hq_cache.get(k)

def hq_cache_set(k, v):
    with _hq_cache_lock: _hq_cache[k] = v

# AI call counters — actual API calls vs cache hits, for cost/throughput accounting.
# verify never caches at this level (uniqueness comes from match_company / name_first
# caches around it). enrich + hq are cached so we count both paths separately.
_ai_calls = {
    "verify_api":    0,  # _openai_verify successful API hits (Tier 3 + name-first)
    "verify_failed": 0,  # _openai_verify exceptions (still costs 0 but logged)
    "enrich_api":    0,  # _openai_enrich API hits
    "enrich_cached": 0,  # _openai_enrich cache hits (no spend)
    "hq_api":        0,  # _openai_hq_search API hits
    "hq_cached":     0,  # _openai_hq_search cache hits (no spend)
    "narrative":     0,  # generate_narrative_summary (no web search; cheaper)
}
_ai_calls_lock = threading.Lock()

def _bump_ai(key: str, n: int = 1):
    with _ai_calls_lock:
        _ai_calls[key] = _ai_calls.get(key, 0) + n

# Per-call cost (rough) — web_search_preview tool dominates. Narrative is plain text.
_AI_COST_USD = {"verify_api": 0.20, "enrich_api": 0.20, "hq_api": 0.20, "narrative": 0.01}

def _ai_calls_total_billable() -> int:
    return _ai_calls["verify_api"] + _ai_calls["enrich_api"] + _ai_calls["hq_api"] + _ai_calls["narrative"]

def _ai_calls_estimated_spend() -> float:
    return sum(_ai_calls.get(k, 0) * c for k, c in _AI_COST_USD.items())

# ═══════════════════════════════════════════════════════════════════════════════
# ██  WRITE QUEUE  (dedicated writer thread — no lock contention)
# ═══════════════════════════════════════════════════════════════════════════════
_wq       = Queue()
_cl       = threading.Lock()    # counter lock
_SENTINEL = object()

ABBREV_MAP = {
    r"\bSt\b":"Street", r"\bBlvd\b":"Boulevard", r"\bRd\b":"Road",
    r"\bAve?\b":"Avenue", r"\bDr\b":"Drive", r"\bLn\b":"Lane",
    r"\bPkwy\b":"Parkway", r"\bCt\b":"Court", r"\bPl\b":"Place",
    r"\bHwy\b":"Highway", r"\bSq\b":"Square", r"\bFwy\b":"Freeway",
    r"\bAv\.\b":"Avenida", r"\bC/\b":"Calle",
}

# ═══════════════════════════════════════════════════════════════════════════════
# ██  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip())

def safe_float(x):
    try:
        s = str(x or "").strip().lower()
        return None if s in ("","nan","none","null") else float(x)
    except: return None

def token_sim(a: str, b: str) -> float:
    a, b = norm(a).lower(), norm(b).lower()
    if not a or not b: return 0.0
    if a == b: return 1.0
    if HAVE_RAPIDFUZZ:
        return rfuzz.token_set_ratio(a, b) / 100.0
    if a in b or b in a: return 0.92
    at = set(re.sub(r"[^\w\s]","",a).split())
    bt = set(re.sub(r"[^\w\s]","",b).split())
    if not at or not bt: return 0.0
    if at.issubset(bt) or bt.issubset(at): return 0.92
    return len(at & bt) / max(len(at), len(bt))

def get_threshold(country: str) -> float:
    return REGION_THRESHOLDS.get(country.upper(), NAME_SIM_THRESHOLD)

def strip_legal(n: str) -> str:
    return re.sub(
        r"\b(llc|inc|corp|ltd|lp|llp|s\.a\.|s\.l\.|gmbh|bv|nv|plc|pty|co\.|company|"
        r"limited|incorporated|corporation|ltda|sas|srl|spa|ag|kg|s\.a\.s\.|s\.r\.l\.|s\.p\.a\.)\b\.?",
        "", n, flags=re.IGNORECASE).strip(" ,.")

def clean_company(n: str) -> str:
    n = re.sub(r"\b(run rate|recap|actuals|budget|forecast|ytd|mtd|qtd|variance|plan|"
               r"baseline|accrual|reforecast|outlook|headcount)\b", "", n, flags=re.IGNORECASE)
    n = re.sub(r"\s+[A-Z]{1,4}\d{3,}\b", "", n, flags=re.IGNORECASE)
    n = re.sub(r"\s+\d{4,}\b", "", n)
    n = re.sub(r"\b(accounts payable|accounts receivable|a/p|a/r|corporate office|"
               r"head office|main office|c/o|attn|procurement|purchasing)\b", "", n, flags=re.IGNORECASE)
    n = re.sub(r"[-–—/]\s*$", "", n)
    n = re.sub(r"\s{2,}", " ", n).strip(" ,.-")
    return n or n

def has_non_latin(t: str) -> bool:
    for ch in (t or ""):
        nm = unicodedata.name(ch, "")
        if any(s in nm for s in ["CJK","ARABIC","CYRILLIC","HEBREW","THAI","HANGUL","HIRAGANA","KATAKANA","DEVANAGARI"]):
            return True
    return False

def detect_lang(t: str) -> str:
    if not HAVE_LANGDETECT or not t or len(t.strip()) < 3: return "en"
    try: return detect(t)
    except: return "en"

def translate_en(t: str) -> str:
    if not HAVE_TRANSLATOR or not t or len(t.strip()) < 3: return t
    try: return GoogleTranslator(source="auto", target="en").translate(t) or t
    except: return t

def normalize_abbrevs(t: str) -> str:
    for pat, rep in ABBREV_MAP.items():
        t = re.sub(pat, rep, t, flags=re.IGNORECASE)
    return t

def clean_field(t) -> str:
    if not isinstance(t, str): return ""
    t = unicodedata.normalize("NFC", t)
    t = re.sub(r"[\x00-\x1f\x7f]", " ", t)
    return re.sub(r"\s+", " ", t).strip()

# Canonical mojibake: a UTF-8 multi-byte lead (U+00C2..U+00EF) decoded as
# Latin-1 leaves a "lead char" followed by a continuation byte (U+0080..U+00BF).
_MOJIBAKE_PAIR_RE = re.compile(r"[Â-ï][-¿]")

def _rescue_mojibake(s: str):
    """Detect and reverse UTF-8 → Latin-1 / cp1252 encoding corruption.
    Returns (possibly_fixed_string, was_rescued).

    Detection: count canonical 'lead+continuation' mojibake pairs. If the
    rescue eliminates all/most of them (and produces valid UTF-8), accept.
    Works for Latin-diacritic (Málaga, Lübecker) AND non-Latin script
    (Chinese, Cyrillic, Arabic, etc.)."""
    if not s or not isinstance(s, str):
        return s, False
    pair_count = len(_MOJIBAKE_PAIR_RE.findall(s))
    if pair_count == 0:
        return s, False
    for src_enc in ("latin-1", "cp1252"):
        try:
            rescued = s.encode(src_enc, errors="strict").decode("utf-8", errors="strict")
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        new_pair_count = len(_MOJIBAKE_PAIR_RE.findall(rescued))
        if new_pair_count < pair_count:
            return rescued, True
    return s, False

def _is_garbage_street(s: str) -> bool:
    """Detect streets that are unusable for geocoding (codes, internal docs, single
    generic words, etc.). City/state/country are kept regardless — only the street
    is dropped when this fires."""
    s = (s or "").strip()
    if not s: return True
    if re.search(r"\b(VAT|EORI|TIN|EIN|GST|TVA|RFC|CUIT|NIT)\s*[A-Z0-9]+", s, re.I): return True
    if re.fullmatch(r"[A-Z0-9\-_/.\s]{1,15}", s): return True
    if s.lower() in {"st","street","road","rd","ave","avenue","main","na","n/a","--","-","none","null"}: return True
    if re.match(r"^(goods receipt|loading dock|warehouse|gate|building|piso|nave|unit|suite|room)\s*\d*\b", s, re.I): return True
    if re.fullmatch(r"[\s\-_/.,]+", s): return True
    return False

def _is_generic_name(name: str) -> bool:
    """Reject company names too generic to web-search productively."""
    n = (name or "").strip().lower()
    if not n or len(n) <= 2: return True
    if n in {"company","corp","corporation","inc","ltd","limited","llc","sa","gmbh",
             "the company","na","n/a","unknown","tbd"}: return True
    if re.fullmatch(r"[\d\s\-_/.,]+", n): return True
    return False

# ═══════════════════════════════════════════════════════════════════════════════
# ██  STEP 1 — ADDRESS PRE-PROCESSING
# ═══════════════════════════════════════════════════════════════════════════════
def preprocess_address(row: pd.Series) -> dict:
    raw   = row.get("FULL_ADDRESS", "")
    parts = [clean_field(p) for p in (raw.split(DELIMITER) if isinstance(raw, str) else [])]
    comp  = {f: (parts[i] if i < len(parts) else "") for i, f in enumerate(ADDR_FIELDS)}

    # Rescue mojibake (UTF-8 mis-decoded as Latin-1/cp1252) before any further
    # processing. Applied per-field so a corrupted street doesn't poison the rest.
    rescued_fields = []
    for f in ADDR_FIELDS:
        fixed, was_rescued = _rescue_mojibake(comp[f])
        if was_rescued:
            comp[f] = fixed
            rescued_fields.append(f)
    mojibake_rescued = bool(rescued_fields)

    sl        = detect_lang(comp["street"])
    nl        = has_non_latin(comp["street"])
    orig      = ""
    translated= False

    if sl != "en" or nl:
        orig             = comp["street"]
        comp["street"]   = translate_en(comp["street"])
        translated       = True

    comp["street"] = normalize_abbrevs(comp["street"])

    addr_for_api = ", ".join(p for p in [
        comp["street"], comp["city"], comp["state"], comp["zip"], comp["country"]
    ] if p)

    flags = []
    if not comp["street"]:  flags.append("MISSING_STREET")
    if not comp["city"]:    flags.append("MISSING_CITY")
    if not comp["country"]: flags.append("MISSING_COUNTRY")
    if comp["street"] and re.fullmatch(r"[A-Z0-9\-_]{1,10}", comp["street"]):
        flags.append("POSSIBLE_PLACEHOLDER_STREET")
    if not comp["zip"] and comp["country"] in ["US","CA","GB","DE","FR"]:
        flags.append("MISSING_ZIP")
    if comp["street"] and _is_garbage_street(comp["street"]):
        flags.append("GARBAGE_STREET")
    if mojibake_rescued:
        flags.append("MOJIBAKE_RESCUED")

    skip = "MISSING_STREET" in flags and "MISSING_CITY" in flags

    return {
        "MDM_KEY":          row.get("MDM_KEY", ""),
        "SOURCE_NAME":      clean_field(str(row.get("SOURCE_NAME", ""))),
        "FULL_ADDRESS_RAW": raw,
        "street":           comp["street"],
        "street_original":  orig,
        "city":             comp["city"],
        "state":            comp["state"],
        "country":          comp["country"],
        "zip":              comp["zip"],
        "address_for_api":  addr_for_api,
        "detected_lang":    sl,
        "has_non_latin":    nl,
        "street_translated":translated,
        "mojibake_rescued": mojibake_rescued,
        "mojibake_fields":  ",".join(rescued_fields),
        "pre_flags":        "|".join(flags) if flags else "OK",
        "skip_api":         skip,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ██  STEP 2 — ADDRESS VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════
def _addr_search(query: str, country: str) -> dict:
    params = {
        "subscription-key": AZURE_MAPS_KEY, "api-version": "1.0",
        "query": query, "limit": 1, "language": "en-US",
    }
    if country and len(country) == 2:
        params["countrySet"] = country.upper()
    try:
        r   = safe_get(ADDRESS_URL, params)
        res = r.json().get("results", [])
        if not res: return {"raw_status": "NO_RESULT", "score": 0.0}
        top  = res[0]
        sc   = top.get("score", 0.0)
        mt   = top.get("type", "")
        addr = top.get("address", {})
        pos  = top.get("position", {})
        if sc >= ADDR_CONFIDENCE_THR and mt in ["Point Address","Address Range","Street"]:
            rs = "VALID"
        elif sc >= ADDR_CONFIDENCE_THR:
            rs = "VALID_LOW_PRECISION"
        elif sc >= 0.3:
            rs = "LOW_CONFIDENCE"
        else:
            rs = "INVALID"
        return {
            "raw_status":        rs, "score": round(sc,4), "match_type": mt,
            "returned_street":   addr.get("streetNameAndNumber", addr.get("streetName","")),
            "returned_city":     addr.get("municipality",""),
            "returned_state":    addr.get("countrySubdivision",""),
            "returned_country":  addr.get("countryCode",""),
            "returned_zip":      addr.get("postalCode",""),
            "returned_freeform": addr.get("freeformAddress",""),
            "lat": pos.get("lat",""), "lon": pos.get("lon",""),
        }
    except Exception as e:
        return {"raw_status": "API_ERROR", "error": str(e), "score": 0.0}

def validate_address(rec: dict) -> dict:
    if rec["skip_api"]:
        return {
            "addr_final_status":"SKIPPED_NO_ADDRESS","addr_correction_strategy":"N/A",
            "val_score":0.0,"val_match_type":"",
            "val_returned_street":"","val_returned_city":"","val_returned_state":"",
            "val_returned_country":"","val_returned_zip":"","val_returned_freeform":"",
            "val_lat":"","val_lon":"",
        }

    result = _addr_search(rec["address_for_api"], rec["country"])
    if result.get("raw_status") not in ("VALID","VALID_LOW_PRECISION"):
        # fallback strategies
        for strategy, query in [
            ("DROP_STREET",    ", ".join(p for p in [rec["city"],rec["state"],rec["zip"],rec["country"]] if p)),
            ("DROP_STATE_ZIP", ", ".join(p for p in [rec["street"],rec["city"],rec["country"]] if p)),
            ("CITY_COUNTRY",   ", ".join(p for p in [rec["city"],rec["country"]] if p)),
        ]:
            if not query.strip(): continue
            r2 = _addr_search(query, rec["country"])
            r2["correction_strategy"] = strategy
            if r2.get("raw_status") in ("VALID","VALID_LOW_PRECISION"):
                result = r2
                break
        else:
            result["correction_strategy"] = result.get("correction_strategy","ALL_FALLBACKS_FAILED")
            result["raw_status"] = result.get("raw_status","MANUAL_REVIEW_NEEDED")

    if ADDR_DELAY: time.sleep(ADDR_DELAY)

    return {
        "addr_final_status":      result.get("raw_status",""),
        "addr_correction_strategy": result.get("correction_strategy","PRIMARY"),
        "val_score":              result.get("score",0.0),
        "val_match_type":         result.get("match_type",""),
        "val_returned_street":    result.get("returned_street",""),
        "val_returned_city":      result.get("returned_city",""),
        "val_returned_state":     result.get("returned_state",""),
        "val_returned_country":   result.get("returned_country",""),
        "val_returned_zip":       result.get("returned_zip",""),
        "val_returned_freeform":  result.get("returned_freeform",""),
        "val_lat":                result.get("lat",""),
        "val_lon":                result.get("lon",""),
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ██  STEP 3 — ROUTE  (FULL_MATCH | NAME_FIRST | SKIP)
# ═══════════════════════════════════════════════════════════════════════════════
ADDR_GATE_SCORE = 0.85   # business floor — do not lower without approval

# Match types where Azure pinned a SPECIFIC building/range — safe for FULL_MATCH.
# "Street" means Azure only matched the street name, not the building number — even
# at score 0.93 this is unsafe (e.g. Bimbo: input "320-80 Avenue Se" missing leading
# "4", Azure dropped the number and returned "80 Avenue Southeast" with a high score).
PRECISE_MATCH_TYPES = {"Point Address", "Address Range"}

def route(rec: dict, addr_result: dict) -> str:
    """Three-way decision after address validation:
      FULL_MATCH  — Azure pinned a specific building (Point Address / Address Range)
                    AND score >= 0.85 AND no fallback strategy was used
      NAME_FIRST  — address unreliable (street-only match, low score, garbage street,
                    fallback strategy used) but name+geo are usable
      SKIP        — neither address nor name usable
    """
    score    = float(addr_result.get("val_score", 0.0) or 0.0)
    status   = str(addr_result.get("addr_final_status", "")).strip().upper()
    strategy = addr_result.get("addr_correction_strategy", "PRIMARY")
    flags    = rec.get("pre_flags", "")
    match_type = str(addr_result.get("val_match_type", "")).strip()

    # Strong address: passes gate AND match_type pins a specific building AND street
    # wasn't fallback-dropped AND not garbage.
    if (score >= ADDR_GATE_SCORE
        and status in ("VALID", "VALID_LOW_PRECISION")
        and strategy == "PRIMARY"
        and "GARBAGE_STREET" not in flags
        and match_type in PRECISE_MATCH_TYPES):
        return "FULL_MATCH"

    # Address is weak/dropped/street-only — try name-first if we have a usable name + some geo
    cleaned  = clean_company(rec.get("SOURCE_NAME", ""))
    has_name = bool(cleaned) and not _is_generic_name(cleaned)
    has_geo  = bool(rec.get("city")) or bool(rec.get("country"))

    if has_name and has_geo:
        return "NAME_FIRST"
    return "SKIP"

def should_send_to_company_match(addr_result: dict) -> bool:
    """Kept for backwards-compatibility; prefer route() in new code."""
    score = float(addr_result.get("val_score", 0.0) or 0.0)
    status = str(addr_result.get("addr_final_status", "")).strip().upper()
    match_type = str(addr_result.get("val_match_type", "")).strip()
    return (score >= ADDR_GATE_SCORE
            and status in ("VALID", "VALID_LOW_PRECISION")
            and match_type in PRECISE_MATCH_TYPES)

# ═══════════════════════════════════════════════════════════════════════════════
# ██  STEP 4 — COMPANY MATCH  (Tier 1 → Tier 2 → Tier 3)
# ═══════════════════════════════════════════════════════════════════════════════
def _geocode(freeform: str, country: str):
    params = {"subscription-key":AZURE_MAPS_KEY,"api-version":"1.0",
              "query":freeform,"limit":1,"language":"en-US"}
    if country and len(country)==2: params["countrySet"]=country.upper()
    try:
        r = safe_get(GEOCODE_URL, params)
        res = r.json().get("results",[])
        if res:
            pos = res[0].get("position",{})
            return pos.get("lat"), pos.get("lon")
    except: pass
    return None, None

def _fuzzy(company, addr, lat, lon, thr) -> dict:
    freeform = addr.get("freeform","")
    country  = addr.get("country","")
    base     = strip_legal(company)
    all_res  = []
    queries  = [
        f"{company} {addr.get('street','')} {addr.get('city','')}",
        f"{company} {addr.get('city','')} {addr.get('state','')}",
        f"{base} {addr.get('city','')} {addr.get('state','')}",
        f"{base} {country}",
    ]
    for q in queries:
        for radius in ([500,2000,20000] if lat and lon else [None]):
            params = {"subscription-key":AZURE_MAPS_KEY,"api-version":"1.0",
                      "query":q,"limit":5,"language":"en-US",
                      "idxSet":"POI,PAD,Addr","maxFuzzyLevel":2}
            if lat and lon and radius: params.update({"lat":lat,"lon":lon,"radius":radius})
            if country and len(country)==2: params["countrySet"]=country.upper()
            try:
                r = safe_get(FUZZY_URL, params)
                res = r.json().get("results",[])
                all_res.extend(res)
                if res: break
            except: pass
    if not all_res: return {"status":"NOT_FOUND"}
    best, bsim = None, -1
    for r in all_res:
        cname = r.get("poi",{}).get("name","") or r.get("address",{}).get("freeformAddress","")
        s = token_sim(company, cname)
        if s > bsim: bsim, best = s, r
    poi = best.get("poi",{}); addr2 = best.get("address",{}); pos = best.get("position",{})
    fa  = norm(addr2.get("freeformAddress",""))
    asim = token_sim(freeform, fa)
    verd = "CORRECT" if asim >= ADDR_SIM_THRESHOLD else ("PARTIAL" if asim >= 0.30 else "WRONG")
    return {"status":"FOUND","name":norm(poi.get("name","")),"name_sim":round(bsim,4),
            "address":fa,"addr_sim":round(asim,4),"addr_verdict":verd,
            "phone":norm(poi.get("phone","")),"url":norm(poi.get("url","")),
            "lat":pos.get("lat",""),"lon":pos.get("lon","")}

def _poi(company, addr, lat, lon, thr) -> dict:
    country = addr.get("country","")
    base    = strip_legal(company)
    all_res = []
    for query in [company, base]:
        for radius in ([500,5000,50000] if lat and lon else [None]):
            params = {"subscription-key":AZURE_MAPS_KEY,"api-version":"1.0",
                      "query":query,"limit":5,"language":"en-US","idxSet":"POI"}
            if lat and lon and radius: params.update({"lat":lat,"lon":lon,"radius":radius})
            if country and len(country)==2: params["countrySet"]=country.upper()
            try:
                r = safe_get(POI_URL, params)
                res = r.json().get("results",[])
                all_res.extend(res)
                if res: break
            except: pass
    if not all_res: return {"status":"NOT_FOUND"}
    best, bsim = None, -1
    for r in all_res:
        s = token_sim(company, r.get("poi",{}).get("name",""))
        if s > bsim: bsim, best = s, r
    return {"status":"FOUND","name":norm(best.get("poi",{}).get("name","")),"name_sim":round(bsim,4),
            "address":norm(best.get("address",{}).get("freeformAddress","")),
            "phone":norm(best.get("poi",{}).get("phone","")),"url":norm(best.get("poi",{}).get("url",""))}

def _occupant(lat, lon, country) -> str:
    if lat is None or lon is None: return ""
    try:
        params = {"subscription-key":AZURE_MAPS_KEY,"api-version":"1.0",
                  "query":"business","limit":3,"language":"en-US",
                  "lat":lat,"lon":lon,"radius":100}
        if country and len(country)==2: params["countrySet"]=country.upper()
        r   = safe_get(POI_URL, params)
        res = r.json().get("results",[])
        nms = [norm(x.get("poi",{}).get("name","")) for x in res if x.get("poi",{}).get("name","")]
        return " | ".join(nms[:3])
    except: return ""

SYSTEM_PROMPT = """You are a senior Master Data Management (MDM) analyst with expertise in:
- Global company name disambiguation across 50+ languages and scripts
- Address validation for multilingual, transliterated, and non-Latin records
- Firmographic data enrichment from authoritative web sources

LANGUAGE & SCRIPT RULES:
1. Company names may be in ANY language (Arabic, Chinese, Russian, Japanese, Korean, Hindi, etc.)
2. Transliterated names (e.g. "Al Faisaliah" = الفيصلية) count as matches
3. Legal suffixes vary by country — match them using local conventions
4. For non-English records, ALWAYS search both the original AND the English translation

VERIFICATION — Chain-of-Thought:
Step 1: Identify the language/script of the company name
Step 2: Generate 3-5 search variants (original, translated, transliterated, abbreviated)
Step 3: Search for each variant + city/country to find official presence
Step 4: Cross-reference found address against MDM address
Step 5: Apply the 3-case decision rule

3-CASE DECISION RULE:
CASE 1 — Company EXISTS at MDM address → company_exists=true, address_match=CORRECT
CASE 2 — Company NOT at MDM address but address is occupied → fill mdm_address_occupant
CASE 3 — Company exists elsewhere → list ALL locations, apply nearest-location rule

ENRICHMENT FIELDS (firmographics):
After verification, populate these company-level fields from authoritative sources
(SEC filings, official company website, D&B, US Census NAICS, OpenCorporates, LinkedIn):

- legal_name: registered legal entity (e.g. "TJX Companies, Inc." for "Tj Maxx 01607")
- parent_company: immediate parent (null if standalone)
- domestic_ultimate: highest entity in same country
- global_ultimate: highest entity worldwide
- is_headquarters: "true"/"false"/"unknown" — is THIS specific MDM address the HQ
- email_domain: primary email domain (derived from website, e.g. "honeywell.com")
- naics_code: 6-digit NAICS code (US Census). MUST be exactly 2/3/4/5/6 digits or null.
- naics_description: human description of the NAICS code
- sic_code: 4-digit SIC code. MUST be exactly 4 digits or null.
- sic_description: human description of the SIC code
- employee_count_range: e.g. "1-10", "11-50", "51-200", "201-500", "501-1000",
  "1001-5000", "5001-10000", "10000+", or null
- revenue_range: e.g. "<$1M", "$1M-$10M", "$10M-$50M", "$50M-$250M", "$250M-$1B",
  "$1B-$10B", "$10B+", or null
- year_established: 4-digit year as string, or null
- industry: short industry label (e.g. "Retail Apparel", "Industrial Conglomerate")

ANTI-HALLUCINATION RULES (STRICT):
- Never invent addresses, websites, phone numbers, NAICS/SIC codes, or any field value
- If uncertain about ANY field, return null for that specific field — do NOT guess
- NAICS codes must be valid US Census codes (digit-format-only validated post-call)
- SIC codes must be exactly 4 digits
- If uncertain about overall verification, set confidence=LOW
- Only report what you found in actual search results — cite sources in evidence
- Empty/null is always preferred over a fabricated value
"""

OUTPUT_SCHEMA = {
    "type":"json_schema","name":"company_verification","schema":{
        "type":"object",
        "properties":{
            "company_exists":             {"type":["boolean","null"]},
            "canonical_name":             {"type":["string","null"]},
            "website":                    {"type":["string","null"]},
            "all_locations":              {"type":"string"},
            "locations_in_state":         {"type":"string"},
            "best_address":               {"type":["string","null"]},
            "real_street":                {"type":["string","null"]},
            "real_city":                  {"type":["string","null"]},
            "real_state":                 {"type":["string","null"]},
            "real_zip":                   {"type":["string","null"]},
            "mdm_address_occupant":       {"type":["string","null"]},
            "nearest_location_reasoning": {"type":["string","null"]},
            "address_match":              {"type":"string","enum":["CORRECT","PARTIAL","WRONG","UNKNOWN"]},
            "confidence":                 {"type":"string","enum":["HIGH","MEDIUM","LOW"]},
            "evidence":                   {"type":"string"},
            # ─── enrichment / firmographics (string|null; null when uncertain) ───
            "legal_name":                 {"type":["string","null"]},
            "parent_company":             {"type":["string","null"]},
            "domestic_ultimate":          {"type":["string","null"]},
            "global_ultimate":            {"type":["string","null"]},
            "is_headquarters":            {"type":["string","null"]},
            "email_domain":               {"type":["string","null"]},
            "naics_code":                 {"type":["string","null"]},
            "naics_description":          {"type":["string","null"]},
            "sic_code":                   {"type":["string","null"]},
            "sic_description":            {"type":["string","null"]},
            "employee_count_range":       {"type":["string","null"]},
            "revenue_range":              {"type":["string","null"]},
            "year_established":           {"type":["string","null"]},
            "industry":                   {"type":["string","null"]},
        },
        "required":["company_exists","canonical_name","website","all_locations",
                    "locations_in_state","best_address","real_street","real_city",
                    "real_state","real_zip","mdm_address_occupant",
                    "nearest_location_reasoning","address_match","confidence","evidence",
                    "legal_name","parent_company","domestic_ultimate","global_ultimate",
                    "is_headquarters","email_domain","naics_code","naics_description",
                    "sic_code","sic_description","employee_count_range","revenue_range",
                    "year_established","industry"],
        "additionalProperties":False
    }
}

# Subset schema for enrichment-only calls (Tier 1/2 paths). No verification fields.
ENRICH_SCHEMA = {
    "type":"json_schema","name":"company_enrichment","schema":{
        "type":"object",
        "properties":{
            "legal_name":                 {"type":["string","null"]},
            "parent_company":             {"type":["string","null"]},
            "domestic_ultimate":          {"type":["string","null"]},
            "global_ultimate":            {"type":["string","null"]},
            "is_headquarters":            {"type":["string","null"]},
            "website":                    {"type":["string","null"]},
            "email_domain":               {"type":["string","null"]},
            "naics_code":                 {"type":["string","null"]},
            "naics_description":          {"type":["string","null"]},
            "sic_code":                   {"type":["string","null"]},
            "sic_description":            {"type":["string","null"]},
            "employee_count_range":       {"type":["string","null"]},
            "revenue_range":              {"type":["string","null"]},
            "year_established":           {"type":["string","null"]},
            "industry":                   {"type":["string","null"]},
            "evidence":                   {"type":"string"},
        },
        "required":["legal_name","parent_company","domestic_ultimate","global_ultimate",
                    "is_headquarters","website","email_domain","naics_code","naics_description",
                    "sic_code","sic_description","employee_count_range","revenue_range",
                    "year_established","industry","evidence"],
        "additionalProperties":False
    }
}

ENRICHMENT_FIELDS = [
    "legal_name","parent_company","domestic_ultimate","global_ultimate",
    "is_headquarters","email_domain","naics_code","naics_description",
    "sic_code","sic_description","employee_count_range","revenue_range",
    "year_established","industry",
]

# Validate NAICS (2-6 digits) and SIC (4 digits). Strip otherwise — model said it's "unsure".
_NAICS_RE = re.compile(r"^\d{2,6}$")
_SIC_RE   = re.compile(r"^\d{4}$")

def _validate_codes(d: dict) -> dict:
    """Wipe NAICS/SIC if they don't match expected digit formats. Anti-hallucination."""
    nc = (d.get("naics_code") or "").strip()
    if nc and not _NAICS_RE.match(nc):
        d["naics_code"] = None
        d["naics_description"] = None
    sc = (d.get("sic_code") or "").strip()
    if sc and not _SIC_RE.match(sc):
        d["sic_code"] = None
        d["sic_description"] = None
    return d

def _enforce_address_match_rigor(ai: dict, mdm_street: str, mdm_city: str = "") -> dict:
    """Deterministic backstop for the verdict-hallucination failure mode where
    the model marks address_match=CORRECT based on state/city overlap without
    actually finding the MDM street in the company's known locations.

    Rule: if address_match=CORRECT, the leading street number from mdm_street
    MUST appear in all_locations. We deliberately DO NOT include best_address
    in the check — the model often parrots the MDM address straight into
    best_address even when it never independently confirmed the location, so
    using best_address would let those exact hallucinations pass.

    all_locations is supposed to be "list ALL locations of the company". If the
    MDM street really is one of them, its street number must appear there.

    Also enforces: when address_match != CORRECT, mdm_address_occupant must be
    non-empty (defaults to a sentinel string so review queue catches it).
    """
    if not isinstance(ai, dict):
        return ai
    match  = (ai.get("address_match") or "").upper()
    street = (mdm_street or "").strip()
    city   = (mdm_city or "").strip().lower()

    if match == "CORRECT" and street:
        nums = re.findall(r"\b\d{2,6}\b", street)
        haystack = (ai.get("all_locations") or "").lower()
        # Require either a matching street number in all_locations, OR (no number
        # in MDM AND city present) — covers cases like "Main Street, London".
        num_ok  = bool(nums) and any(n in haystack for n in nums)
        city_ok = (not nums) and city and city in haystack
        if not (num_ok or city_ok):
            ai["address_match"] = "PARTIAL"
            ai["confidence"]    = "LOW"
            note = (f" [PIPELINE DOWNGRADE: MDM street '{street}' not present in "
                    f"all_locations — address_match set to PARTIAL because state/"
                    f"city-level overlap (or best_address parroting) is insufficient "
                    f"evidence of an actual location at the MDM address.]")
            ai["evidence"] = (ai.get("evidence", "") or "") + note
            match = "PARTIAL"

    if match != "CORRECT":
        occ = (ai.get("mdm_address_occupant") or "").strip()
        if not occ:
            ai["mdm_address_occupant"] = "no business indexed at this address"
    return ai

def _extract_sources(rsp) -> List[dict]:
    """Pull url_citation annotations from a Responses API output. Each is a URL the
    model actually grounded against via web_search_preview — ground truth, not
    self-reported. Deduped by URL within the call."""
    seen = {}
    for item in (getattr(rsp, "output", None) or []):
        for content in (getattr(item, "content", None) or []):
            for ann in (getattr(content, "annotations", None) or []):
                if getattr(ann, "type", "") == "url_citation":
                    url = (getattr(ann, "url", "") or "").strip()
                    if url and url not in seen:
                        seen[url] = {"url": url, "title": (getattr(ann, "title", "") or "").strip()}
    return list(seen.values())

def _dedup_sources(srcs: List[dict]) -> List[dict]:
    """Dedup a list of {url, title} dicts by URL, preserving order."""
    seen = {}
    for s in (srcs or []):
        url = (s.get("url") or "").strip()
        if url and url not in seen:
            seen[url] = s
    return list(seen.values())

def _format_sources(srcs: List[dict]) -> str:
    """Pipe-separated 'title (url)' for the CSV sources column."""
    out = []
    for s in (srcs or []):
        url = (s.get("url") or "").strip()
        if not url: continue
        title = (s.get("title") or "").strip()
        out.append(f"{title} ({url})" if title else url)
    return " | ".join(out)

def _openai_verify(company, cleaned, addr, lat, lon, azure_ev):
    prompt = f"""
<mdm_record>
  <original_name>{company}</original_name>
  <cleaned_name>{cleaned}</cleaned_name>
  <street>{addr.get("street","")}</street>
  <city>{addr.get("city","")}</city>
  <state>{addr.get("state","")}</state>
  <zip>{addr.get("zip","")}</zip>
  <country>{addr.get("country","")}</country>
  <full_address>{addr.get("freeform","")}</full_address>
  <coordinates>lat={lat}, lon={lon}</coordinates>
</mdm_record>
<azure_pre_search>{json.dumps(azure_ev, indent=2)}</azure_pre_search>
<execute_these_searches_in_order>
SEARCH 1: "{cleaned}" + "{addr.get("city","")}" "{addr.get("country","")}"
SEARCH 2: "{cleaned}" official website OR headquarters
SEARCH 3: "{addr.get("street","")}" "{addr.get("city","")}" — what business is at this address?
SEARCH 4: If not found — try translated/transliterated name variants
SEARCH 5 (enrichment): "{cleaned}" NAICS code SIC code — find from US Census or SEC filings
SEARCH 6 (enrichment): "{cleaned}" parent company OR ultimate parent OR corporate hierarchy
SEARCH 7 (enrichment): "{cleaned}" employees revenue year founded — from official sources
NEAREST LOCATION RULE: same state as "{addr.get("state","")}" → always pick first.
OUTPUT: Fill all JSON fields including verification AND enrichment.
        For ANY enrichment field where you are uncertain → return null. Never fabricate.
        Cite sources in evidence.
</execute_these_searches_in_order>"""
    try:
        rsp = oai_client.responses.create(
            model=MODEL_NAME,
            input=[{"role":"system","content":SYSTEM_PROMPT},{"role":"user","content":prompt}],
            tools=[{"type":"web_search_preview"}],
            tool_choice="required",
            text={"format":OUTPUT_SCHEMA},
        )
        _bump_ai("verify_api")
        parsed = _validate_codes(json.loads(rsp.output_text))
        parsed = _enforce_address_match_rigor(parsed,
                                              mdm_street=addr.get("street",""),
                                              mdm_city=addr.get("city",""))
        return parsed, _extract_sources(rsp)
    except Exception as e:
        _bump_ai("verify_failed")
        logging.error(f"OpenAI failed [{company}]: {e}")
        return {"company_exists":None,"canonical_name":None,"website":None,
                "all_locations":"","locations_in_state":"","best_address":None,
                "real_street":None,"real_city":None,"real_state":None,"real_zip":None,
                "mdm_address_occupant":None,"nearest_location_reasoning":None,
                "address_match":"UNKNOWN","confidence":"LOW","evidence":f"OpenAI error: {e}",
                "legal_name":None,"parent_company":None,"domestic_ultimate":None,
                "global_ultimate":None,"is_headquarters":None,"email_domain":None,
                "naics_code":None,"naics_description":None,"sic_code":None,
                "sic_description":None,"employee_count_range":None,"revenue_range":None,
                "year_established":None,"industry":None}, []

def _empty_enrich() -> dict:
    return {f: None for f in ENRICHMENT_FIELDS} | {"website": None, "evidence": ""}

# Map AI's HIGH/MEDIUM/LOW labels to numeric scores for the confidence column
_AI_CONF_SCORE = {"HIGH": 0.90, "MEDIUM": 0.65, "LOW": 0.30}

def _ai_score(label: str) -> float:
    return _AI_CONF_SCORE.get((label or "").strip().upper(), 0.0)

def _openai_enrich(canonical_name: str, country: str, mdm_city: str = ""):
    """Firmographic enrichment for a company that's already been verified by Azure
    (Tier 1 / Tier 2 hit). No address verification — Azure already confirmed location.
    Caches by (canonical_name.lower(), country.upper()): one OpenAI call per unique
    company across the entire run.

    Returns (result_dict, sources_list).
    """
    name = norm(canonical_name)
    if not name or _is_generic_name(name):
        return _empty_enrich() | {"evidence": "skipped: name empty or too generic"}, []

    key = (name.lower(), (country or "").upper())
    cached = enrich_cache_get(key)
    if cached is not None:
        _bump_ai("enrich_cached")
        result_c, sources_c = cached
        return dict(result_c), list(sources_c)

    prompt = f"""
<company>
  <name>{name}</name>
  <country>{country}</country>
  <reference_city>{mdm_city}</reference_city>
</company>
<execute_these_searches>
SEARCH 1: "{name}" official website OR headquarters {country}
SEARCH 2: "{name}" NAICS code SIC code — from US Census, SEC EDGAR, or company filings
SEARCH 3: "{name}" parent company OR ultimate parent OR corporate hierarchy
SEARCH 4: "{name}" employees revenue year founded
OUTPUT: Fill all JSON fields. For ANY field where you are uncertain → return null.
        NAICS must be 2-6 digits exact. SIC must be 4 digits exact. Otherwise null.
        Never fabricate. Empty/null is preferred over a guess. Cite sources in evidence.
</execute_these_searches>"""
    sources: List[dict] = []
    try:
        rsp = oai_client.responses.create(
            model=MODEL_NAME,
            input=[{"role":"system","content":SYSTEM_PROMPT},{"role":"user","content":prompt}],
            tools=[{"type":"web_search_preview"}],
            tool_choice="required",
            text={"format":ENRICH_SCHEMA},
        )
        _bump_ai("enrich_api")
        result = _validate_codes(json.loads(rsp.output_text))
        sources = _extract_sources(rsp)
    except Exception as e:
        logging.error(f"OpenAI enrich failed [{name}]: {e}")
        result = _empty_enrich() | {"evidence": f"OpenAI enrich error: {e}"}

    enrich_cache_set(key, (result, sources))
    if COMPANY_DELAY: time.sleep(COMPANY_DELAY)
    return result, sources

def _empty_company_result(status: str = "", evidence: str = "") -> dict:
    """All company-match output keys, populated empty. Used by skip paths and as
    the base for both match_company and match_company_name_first."""
    return {
        "match_status": status, "company_exists": "", "tier_used": "",
        "canonical_name": "", "website": "", "all_known_locations": "",
        "locations_in_state": "", "best_address": "", "best_street": "",
        "best_city": "", "best_state": "", "best_zip": "", "mdm_address_occupant": "",
        "nearest_location_reasoning": "", "address_match": "",
        "ai_confidence": "", "ai_evidence": evidence,
        "fuzzy_match_name": "", "fuzzy_similarity": "", "fuzzy_found_address": "",
        "poi_match_name": "", "poi_similarity": "",
        "confidence_score": "", "confidence_reason": "",
        # enrichment defaults — empty until _openai_enrich or Tier 3 fills them
        "legal_name": "", "parent_company": "", "domestic_ultimate": "",
        "global_ultimate": "", "is_headquarters": "", "email_domain": "",
        "naics_code": "", "naics_description": "", "sic_code": "",
        "sic_description": "", "employee_count_range": "", "revenue_range": "",
        "year_established": "", "industry": "",
        "enrich_source": "",
        "headquarters_address": "", "headquarters_source": "",
        # internal: raw source lists per call site (formatted into CSV columns
        # by process_record). DictWriter ignores extra keys via extrasaction.
        "_verify_sources": [], "_enrich_sources": [], "_hq_sources": [],
    }

def _ai_to_result_fields(ai: dict, occupant: str, tier: str) -> dict:
    """Translate OpenAI verify response into the company-match output schema."""
    exists = ai.get("company_exists")
    match  = ai.get("address_match", "UNKNOWN")
    if   exists is True  and match == "CORRECT":            status = "COMPANY_FOUND"
    elif exists is True  and match in ("PARTIAL","WRONG"):  status = "COMPANY_FOUND_DIFF_ADDR"
    elif exists is False:                                   status = "COMPANY_NOT_FOUND"
    else:                                                   status = "UNVERIFIED"
    return {
        "match_status":               status,
        "company_exists":             exists,
        "tier_used":                  tier,
        "canonical_name":             ai.get("canonical_name","") or "",
        "website":                    ai.get("website","") or "",
        "all_known_locations":        ai.get("all_locations","") or "",
        "locations_in_state":         ai.get("locations_in_state","") or "",
        "best_address":               ai.get("best_address","") or "",
        "best_street":                ai.get("real_street","") or "",
        "best_city":                  ai.get("real_city","") or "",
        "best_state":                 ai.get("real_state","") or "",
        "best_zip":                   ai.get("real_zip","") or "",
        "mdm_address_occupant":       (ai.get("mdm_address_occupant","") or "") or occupant,
        "nearest_location_reasoning": ai.get("nearest_location_reasoning","") or "",
        "address_match":              match,
        "ai_confidence":              ai.get("confidence","") or "",
        "ai_evidence":                ai.get("evidence","") or "",
        # enrichment fields (returned by extended Tier 3 schema)
        "legal_name":                 ai.get("legal_name","") or "",
        "parent_company":             ai.get("parent_company","") or "",
        "domestic_ultimate":          ai.get("domestic_ultimate","") or "",
        "global_ultimate":            ai.get("global_ultimate","") or "",
        "is_headquarters":            ai.get("is_headquarters","") or "",
        "email_domain":               ai.get("email_domain","") or "",
        "naics_code":                 ai.get("naics_code","") or "",
        "naics_description":          ai.get("naics_description","") or "",
        "sic_code":                   ai.get("sic_code","") or "",
        "sic_description":            ai.get("sic_description","") or "",
        "employee_count_range":       ai.get("employee_count_range","") or "",
        "revenue_range":              ai.get("revenue_range","") or "",
        "year_established":           ai.get("year_established","") or "",
        "industry":                   ai.get("industry","") or "",
        "enrich_source":              tier,
    }

def _apply_enrich(base: dict, enrich: dict, source: str) -> dict:
    """Merge an enrichment dict (from _openai_enrich or _enrich_cache) onto a base
    company-match result. Only fills enrichment fields — does not touch verification.
    """
    for f in ENRICHMENT_FIELDS:
        v = enrich.get(f)
        base[f] = v if v not in (None, "") else base.get(f, "")
    # website is shared between verification and enrichment — only fill if base is empty
    if not base.get("website") and enrich.get("website"):
        base["website"] = enrich["website"]
    base["enrich_source"] = source
    return base

# ═══════════════════════════════════════════════════════════════════════════════
# ██  HEADQUARTERS RESOLUTION
# ═══════════════════════════════════════════════════════════════════════════════
HQ_SCHEMA = {
    "type":"json_schema","name":"headquarters_lookup","schema":{
        "type":"object",
        "properties":{
            "headquarters_address": {"type":["string","null"]},
            "headquarters_city":    {"type":["string","null"]},
            "headquarters_state":   {"type":["string","null"]},
            "headquarters_country": {"type":["string","null"]},
            "source_used":          {"type":["string","null"]},
            "evidence":             {"type":"string"},
        },
        "required":["headquarters_address","headquarters_city","headquarters_state",
                    "headquarters_country","source_used","evidence"],
        "additionalProperties":False
    }
}

HQ_SYSTEM_PROMPT = """You are an MDM analyst resolving the corporate headquarters address
for a company entity. The target is the highest-known entity in its corporate hierarchy
(global ultimate > domestic ultimate > parent > the company itself).

SEARCH PRIORITY ORDER:
1. The company's official website (especially its Contact / About / Locations / Investor pages)
2. SEC EDGAR filings (10-K, 10-Q list registered HQ on the cover page)
3. OpenCorporates, D&B, Wikipedia infoboxes
4. General web search for "<entity> headquarters" or "<entity> head office address"

RULES:
- If multiple HQs are listed (e.g. "global HQ" vs "EMEA HQ"), prefer the GLOBAL ultimate's HQ
- For privately-held companies without public HQ data, return null — do NOT guess
- Return a single canonical street address; do not concatenate multiple offices
- Cite which source you found it on (e.g. "company_website", "sec_edgar", "wikipedia")
- If uncertain, return null. Empty/null is preferred over a fabricated value.
"""

def _openai_hq_search(target_name: str, country: str, website: str = ""):
    """Find the headquarters address for `target_name` (parent/ultimate hierarchy entity).
    Cached by (target_name.lower(), country.upper()): each unique target costs one call.
    Falls back to website search if direct queries fail.

    Returns (result_dict, sources_list).
    """
    target = norm(target_name)
    if not target or _is_generic_name(target):
        return ({"headquarters_address": None, "source_used": "skipped_generic_name",
                 "evidence": "Target name empty or too generic"}, [])

    key = (target.lower(), (country or "").upper())
    cached = hq_cache_get(key)
    if cached is not None:
        _bump_ai("hq_cached")
        result_c, sources_c = cached
        return dict(result_c), list(sources_c)

    site_clause = f'site:{website}' if website else 'official website'
    prompt = f"""
<target>
  <entity_name>{target}</entity_name>
  <country>{country}</country>
  <known_website>{website or "unknown"}</known_website>
</target>
<execute_searches_in_order>
SEARCH 1: "{target}" headquarters address {country}
SEARCH 2: "{target}" "head office" OR "registered office" OR "principal office"
SEARCH 3: {site_clause} contact OR about OR locations OR "investor relations"
SEARCH 4: SEC EDGAR 10-K cover page for "{target}" — find Item 1 / address of principal executive offices
SEARCH 5 (last resort): wikipedia "{target}" infobox — headquarters field
OUTPUT: One canonical headquarters street address, with city/state/country split out.
        Cite the source you used (company_website, sec_edgar, wikipedia, etc.).
        If you cannot verify a single HQ, return null for the address — do not guess.
</execute_searches_in_order>"""
    sources: List[dict] = []
    try:
        rsp = oai_client.responses.create(
            model=MODEL_NAME,
            input=[{"role":"system","content":HQ_SYSTEM_PROMPT},{"role":"user","content":prompt}],
            tools=[{"type":"web_search_preview"}],
            tool_choice="required",
            text={"format":HQ_SCHEMA},
        )
        _bump_ai("hq_api")
        result = json.loads(rsp.output_text)
        sources = _extract_sources(rsp)
    except Exception as e:
        logging.error(f"HQ search failed [{target}]: {e}")
        result = {"headquarters_address": None, "headquarters_city": None,
                  "headquarters_state": None, "headquarters_country": None,
                  "source_used": "error", "evidence": f"OpenAI HQ search error: {e}"}

    hq_cache_set(key, (result, sources))
    if COMPANY_DELAY: time.sleep(COMPANY_DELAY)
    return result, sources

def _format_hq_address(hq: dict) -> str:
    """Compose a single-line address from the HQ search response."""
    parts = [
        (hq.get("headquarters_address") or "").strip(),
        (hq.get("headquarters_city")    or "").strip(),
        (hq.get("headquarters_state")   or "").strip(),
        (hq.get("headquarters_country") or "").strip(),
    ]
    return ", ".join(p for p in parts if p)

def _resolve_hq(rec: dict, comp_res: dict, addr_result: Optional[dict] = None) -> dict:
    """Determine the corporate headquarters address for the matched company.

    Logic:
      1. If is_headquarters=true → the verified MDM address IS the HQ. Done.
      2. Else, climb the hierarchy: global_ultimate > domestic_ultimate > parent_company
         > canonical_name. Take the highest non-empty entity as the HQ search target.
      3. Fire _openai_hq_search(target, country, website). The search itself includes
         a website-fallback step.
      4. If still no result, return source=NOT_FOUND.

    Returns: {"headquarters_address": str, "headquarters_source": str}
    """
    addr_result = addr_result or {}
    # Path 1 — this MDM address is already the HQ
    is_hq = str(comp_res.get("is_headquarters", "") or "").strip().lower()
    if is_hq == "true":
        verified = (addr_result.get("val_returned_freeform", "")
                    or comp_res.get("best_address", "")
                    or rec.get("address_for_api", "")).strip()
        if verified:
            return {"headquarters_address": verified, "headquarters_source": "MDM_VERIFIED_HQ",
                    "_hq_sources": []}

    # Path 2 — climb hierarchy and search
    target = (
        (comp_res.get("global_ultimate", "") or "").strip()
        or (comp_res.get("domestic_ultimate", "") or "").strip()
        or (comp_res.get("parent_company", "") or "").strip()
        or (comp_res.get("canonical_name", "") or "").strip()
    )
    if not target:
        return {"headquarters_address": "", "headquarters_source": "NO_HIERARCHY_TARGET",
                "_hq_sources": []}

    country = rec.get("country", "")
    website = (comp_res.get("website", "") or "").strip()

    hq, hq_sources = _openai_hq_search(target, country, website)
    address = _format_hq_address(hq)
    src     = (hq.get("source_used") or "").strip()

    if address:
        return {
            "headquarters_address": address,
            "headquarters_source":  f"HIERARCHY_LOOKUP:{target} via {src or 'web_search'}",
            "_hq_sources":          hq_sources,
        }
    return {
        "headquarters_address": "",
        "headquarters_source":  f"NOT_FOUND:{target}",
        "_hq_sources":          hq_sources,
    }

def match_company(rec: dict, addr_result: dict) -> dict:
    """Run Tier1 → Tier2 → Tier3 company match on a pre-validated record."""
    company  = norm(rec.get("SOURCE_NAME",""))
    cleaned  = clean_company(company)
    country  = rec.get("country","")
    freeform = addr_result.get("val_returned_freeform","") or rec.get("address_for_api","")
    thr      = get_threshold(country)

    addr = {
        "street":  rec.get("street",""),
        "city":    rec.get("city",""),
        "state":   rec.get("state",""),
        "country": country,
        "zip":     rec.get("zip",""),
        "freeform":freeform,
    }

    ck = (cleaned.lower(), freeform.lower())
    cached = cache_get(ck)
    if cached:
        r = dict(cached)
        r["match_status"] = r.get("match_status","") + "_CACHED"
        return r

    lat = safe_float(addr_result.get("val_lat")) 
    lon = safe_float(addr_result.get("val_lon"))
    if lat is None or lon is None:
        lat, lon = _geocode(freeform, country)

    base = _empty_company_result()

    # Tier 1
    fz = _fuzzy(company, addr, lat, lon, thr)
    base["fuzzy_match_name"]    = fz.get("name","")
    base["fuzzy_similarity"]    = fz.get("name_sim","")
    base["fuzzy_found_address"] = fz.get("address","")
    if fz.get("status")=="FOUND" and fz.get("name_sim",0)>=thr and fz.get("addr_verdict")=="CORRECT":
        sim = float(fz.get("name_sim",0) or 0)
        asim = float(fz.get("addr_sim",0) or 0)
        base.update({"match_status":"COMPANY_FOUND","company_exists":True,"tier_used":"TIER1_FUZZY",
                     "canonical_name":fz["name"],"best_address":fz["address"],
                     "address_match":"CORRECT","ai_confidence":"HIGH",
                     "confidence_score": round(sim, 3),
                     "confidence_reason": (
                         f"Tier 1 (Azure Fuzzy POI): name similarity {sim:.2f} "
                         f">= threshold {thr:.2f}; address verdict CORRECT (addr_sim={asim:.2f}). "
                         f"Azure returned a high-similarity company at the validated address."
                     )})
        # Enrichment for Tier 1 hit — cached by (canonical_name, country)
        enrich, enrich_sources = _openai_enrich(fz["name"], country, addr.get("city",""))
        _apply_enrich(base, enrich, "TIER1_ENRICHED")
        base["_enrich_sources"] = enrich_sources
        # HQ resolution: this address verified, so if is_headquarters=true we use it,
        # else we climb the hierarchy to find the parent/ultimate's HQ.
        base.update(_resolve_hq(rec, base, addr_result))
        cache_set(ck, base); return base

    # Tier 2
    po = _poi(company, addr, lat, lon, thr)
    base["poi_match_name"]  = po.get("name","")
    base["poi_similarity"]  = po.get("name_sim","")
    if po.get("status")=="FOUND" and po.get("name_sim",0)>=thr:
        sim = float(po.get("name_sim",0) or 0)
        # POI lacks an address verdict — downweight to reflect partial verification
        score = round(min(sim * 0.85, 0.85), 3)
        base.update({"match_status":"COMPANY_FOUND","company_exists":True,"tier_used":"TIER2_POI",
                     "canonical_name":po["name"],"best_address":po.get("address",""),
                     "address_match":"PARTIAL","ai_confidence":"MEDIUM",
                     "confidence_score": score,
                     "confidence_reason": (
                         f"Tier 2 (Azure POI search): name similarity {sim:.2f} "
                         f">= threshold {thr:.2f}, but address was not independently re-verified "
                         f"by POI — score downweighted to {score:.2f}."
                     )})
        # Enrichment for Tier 2 hit — cached by (canonical_name, country)
        enrich, enrich_sources = _openai_enrich(po["name"], country, addr.get("city",""))
        _apply_enrich(base, enrich, "TIER2_ENRICHED")
        base["_enrich_sources"] = enrich_sources
        # HQ resolution
        base.update(_resolve_hq(rec, base, addr_result))
        cache_set(ck, base); return base

    # Tier 3 — skip if no address AND no Azure signal
    if fz.get("name_sim",0) < 0.5 and po.get("name_sim",0) < 0.5 and not freeform:
        fz_sim = float(fz.get("name_sim",0) or 0)
        po_sim = float(po.get("name_sim",0) or 0)
        base.update({"match_status":"COMPANY_NOT_FOUND","company_exists":False,
                     "tier_used":"SKIPPED_NO_ADDR","ai_confidence":"LOW",
                     "ai_evidence":"No address + no Azure match — OpenAI skipped",
                     "confidence_score": 0.0,
                     "confidence_reason": (
                         f"NOT FOUND — Tier 1 fuzzy similarity {fz_sim:.2f} and Tier 2 POI "
                         f"similarity {po_sim:.2f} both below 0.5, and no usable freeform "
                         f"address available. OpenAI web search was skipped to avoid wasted "
                         f"calls on unmatchable records."
                     )})
        cache_set(ck, base); return base

    occ = _occupant(lat, lon, country)
    ai, verify_sources = _openai_verify(company, cleaned, addr, lat, lon,
                         {"fuzzy":fz,"poi":po,"occupant":occ,"coords":{"lat":lat,"lon":lon}})
    base.update(_ai_to_result_fields(ai, occ, "TIER3_AI"))
    base["_verify_sources"] = verify_sources
    # Build confidence reason for Tier 3 (always, regardless of found/not-found)
    ai_conf = (ai.get("confidence") or "").upper()
    base["confidence_score"]  = round(_ai_score(ai_conf), 3)
    fz_sim = float(fz.get("name_sim",0) or 0)
    po_sim = float(po.get("name_sim",0) or 0)
    if base.get("company_exists") is True:
        base["confidence_reason"] = (
            f"Tier 3 (OpenAI web search) — confidence={ai_conf}; "
            f"address_match={base.get('address_match','UNKNOWN')}. "
            f"Reached after Tier 1 ({fz_sim:.2f}) and Tier 2 ({po_sim:.2f}) "
            f"failed to verify. Evidence: {base.get('ai_evidence','')[:300]}"
        )
    else:
        base["confidence_reason"] = (
            f"NOT FOUND via Tier 3 — confidence={ai_conf}. "
            f"Tier 1 score {fz_sim:.2f}, Tier 2 score {po_sim:.2f}; "
            f"OpenAI web search did not confirm the company at this address. "
            f"Evidence: {base.get('ai_evidence','')[:300]}"
        )
    # Populate _enrich_cache from this Tier 3 call so future Tier 1/2 hits for same
    # company hit the cache. Only when company verified — don't cache failures.
    canon = base.get("canonical_name","")
    if canon and base.get("company_exists") is True:
        enrich_cache_set(
            (canon.lower(), (country or "").upper()),
            ({f: ai.get(f) for f in ENRICHMENT_FIELDS}
             | {"website": ai.get("website"), "evidence": ai.get("evidence","")},
             verify_sources)
        )
    # HQ resolution — only attempt for verified companies
    if base.get("company_exists") is True:
        base.update(_resolve_hq(rec, base, addr_result))
    cache_set(ck, base)
    if COMPANY_DELAY: time.sleep(COMPANY_DELAY)
    return base

# ═══════════════════════════════════════════════════════════════════════════════
# ██  STEP 4b — NAME-FIRST MATCH (when address fails the 0.85 gate)
# ═══════════════════════════════════════════════════════════════════════════════
def match_company_name_first(rec: dict) -> dict:
    """Address is unusable (low confidence or garbage street). Search by company
    name + whatever geographic info we have (city/state/country/zip), and let
    OpenAI find the canonical address. Skips Azure POI/Fuzzy entirely."""
    company = norm(rec.get("SOURCE_NAME", ""))
    cleaned = clean_company(company)
    country = rec.get("country", "")
    city    = rec.get("city", "")
    state   = rec.get("state", "")
    zipc    = rec.get("zip", "")
    flags   = rec.get("pre_flags", "")

    if not cleaned or _is_generic_name(cleaned):
        r = _empty_company_result(
            status="SKIPPED_NO_USABLE_NAME",
            evidence="Cleaned name empty or too generic for web search",
        )
        r["confidence_score"]  = 0.0
        r["confidence_reason"] = (
            "SKIPPED — cleaned company name is empty or too generic "
            "(e.g. 'corp', 'na', '--', short numeric strings) to web-search "
            "productively. Address was also below the 0.85 confidence gate, "
            "so name-first fallback could not run."
        )
        return r

    # Cache by (cleaned_name, country) — same company in same country reuses lookup
    nk = (cleaned.lower(), country.upper())
    cached = name_cache_get(nk)
    if cached:
        r = dict(cached)
        r["match_status"] = r.get("match_status", "") + "_NAME_CACHED"
        return r

    # Drop ONLY the street if garbage; keep city/state/country/zip
    street = "" if "GARBAGE_STREET" in flags else rec.get("street", "")
    parts    = [p for p in [street, city, state, zipc, country] if p]
    freeform = ", ".join(parts)

    addr = {
        "street":   street,
        "city":     city,
        "state":    state,
        "country":  country,
        "zip":      zipc,
        "freeform": freeform,
    }

    azure_ev = {
        "mode":   "name_first",
        "note":   "Original street was unusable (low Azure confidence or GARBAGE_STREET).",
        "kept":   {"city": city, "state": state, "country": country, "zip": zipc},
        "dropped": {"street_was_garbage": "GARBAGE_STREET" in flags},
    }
    ai, verify_sources = _openai_verify(company, cleaned, addr, None, None, azure_ev)

    result = _empty_company_result()
    result.update(_ai_to_result_fields(ai, "", "TIER3_NAME_FIRST"))
    result["_verify_sources"] = verify_sources
    ai_conf = (ai.get("confidence") or "").upper()
    result["confidence_score"]  = round(_ai_score(ai_conf), 3)
    if result.get("company_exists") is True:
        result["confidence_reason"] = (
            f"Tier 3 NAME-FIRST (OpenAI web search) — confidence={ai_conf}; "
            f"address_match={result.get('address_match','UNKNOWN')}. "
            f"Original street was unusable so the search keyed on company name + "
            f"city/state/country. Evidence: {result.get('ai_evidence','')[:300]}"
        )
    else:
        result["confidence_reason"] = (
            f"NOT FOUND via Tier 3 NAME-FIRST — confidence={ai_conf}. "
            f"OpenAI web search using name + city/state/country did not find a "
            f"verifiable match. Evidence: {result.get('ai_evidence','')[:300]}"
        )
    # HQ resolution — only attempt for verified companies. addr_result is empty for
    # name-first since Azure address validation didn't pass; _resolve_hq falls back
    # to comp_res["best_address"] (set by OpenAI verify) when needed.
    if result.get("company_exists") is True:
        result.update(_resolve_hq(rec, result, addr_result=None))
    name_cache_set(nk, result)
    if COMPANY_DELAY: time.sleep(COMPANY_DELAY)
    return result

# ═══════════════════════════════════════════════════════════════════════════════
# ██  WORKER NODE  — all 4 steps for ONE record
# ═══════════════════════════════════════════════════════════════════════════════
def process_record(row: pd.Series) -> dict:
    t0        = time.time()
    worker_id = threading.current_thread().name   # e.g. "mdm_0", "mdm_1" ...

    # ── Step 1: pre-process ──────────────────────────────────────────────────
    rec = preprocess_address(row)

    # ── Step 2: validate address ─────────────────────────────────────────────
    addr_res = validate_address(rec)

    # ── Step 3: route ────────────────────────────────────────────────────────
    decision = route(rec, addr_res)
    send     = (decision == "FULL_MATCH")

    # ── Step 4: company match (path depends on routing decision) ─────────────
    if decision == "FULL_MATCH":
        comp_res = match_company(rec, addr_res)
    elif decision == "NAME_FIRST":
        comp_res = match_company_name_first(rec)
    else:  # SKIP
        score = float(addr_res.get("val_score", 0.0) or 0.0)
        cleaned = clean_company(rec.get("SOURCE_NAME", ""))
        name_unusable = (not cleaned) or _is_generic_name(cleaned)
        comp_res = _empty_company_result(
            status="SKIPPED_NO_USABLE_DATA",
            evidence=f"Address below {ADDR_GATE_SCORE} confidence and name unusable for web search",
        )
        comp_res["confidence_score"]  = 0.0
        comp_res["confidence_reason"] = (
            f"SKIPPED — address validation score {score:.2f} below the {ADDR_GATE_SCORE} "
            f"gate AND company name "
            + ("is empty/too generic" if name_unusable else "+ city/country fallback unavailable")
            + f". Pre-processing flags: {rec.get('pre_flags','OK')}. "
            f"No reliable signal to route to either company match path."
        )

    elapsed = round(time.time() - t0, 2)

    # ── Format source URLs (raw lists → CSV-friendly strings) ─────────────────
    verify_srcs = comp_res.get("_verify_sources", []) or []
    enrich_srcs = comp_res.get("_enrich_sources", []) or []
    hq_srcs     = comp_res.get("_hq_sources", []) or []
    combined    = _dedup_sources(verify_srcs + enrich_srcs + hq_srcs)

    # ── Merge everything into one flat record ─────────────────────────────────
    out = {
        "MDM_KEY":          rec["MDM_KEY"],
        "SOURCE_NAME":      rec["SOURCE_NAME"],
        "FULL_ADDRESS_RAW": rec["FULL_ADDRESS_RAW"],
        # step 1
        "street":           rec["street"],
        "street_original":  rec["street_original"],
        "city":             rec["city"],
        "state":            rec["state"],
        "country":          rec["country"],
        "zip":              rec["zip"],
        "address_for_api":  rec["address_for_api"],
        "detected_lang":    rec["detected_lang"],
        "has_non_latin":    rec["has_non_latin"],
        "street_translated":rec["street_translated"],
        "mojibake_rescued": rec["mojibake_rescued"],
        "mojibake_fields":  rec["mojibake_fields"],
        "pre_flags":        rec["pre_flags"],
        "skip_api":         rec["skip_api"],
        # step 2
        **{f"val_{k}" if not k.startswith("val_") and not k.startswith("addr_") else k: v
           for k, v in addr_res.items()},
        # step 3
        "routing_decision":      decision,
        "sent_to_company_match": send,
        # step 4
        "match_status":               comp_res["match_status"],
        "company_exists":             comp_res["company_exists"],
        "tier_used":                  comp_res["tier_used"],
        "canonical_name":             comp_res["canonical_name"],
        "website":                    comp_res["website"],
        "all_known_locations":        comp_res["all_known_locations"],
        "locations_in_state":         comp_res["locations_in_state"],
        "best_address":               comp_res["best_address"],
        "best_street":                comp_res["best_street"],
        "best_city":                  comp_res["best_city"],
        "best_state":                 comp_res["best_state"],
        "best_zip":                   comp_res["best_zip"],
        "mdm_address_occupant":       comp_res["mdm_address_occupant"],
        "nearest_location_reasoning": comp_res["nearest_location_reasoning"],
        "address_match":              comp_res["address_match"],
        "ai_confidence":              comp_res["ai_confidence"],
        "ai_evidence":                comp_res["ai_evidence"],
        "fuzzy_match_name":           comp_res["fuzzy_match_name"],
        "fuzzy_similarity":           comp_res["fuzzy_similarity"],
        "fuzzy_found_address":        comp_res["fuzzy_found_address"],
        "poi_match_name":             comp_res["poi_match_name"],
        "poi_similarity":             comp_res["poi_similarity"],
        # per-record confidence
        "confidence_score":           comp_res.get("confidence_score",""),
        "confidence_reason":          comp_res.get("confidence_reason",""),
        # enrichment
        "legal_name":                 comp_res.get("legal_name",""),
        "parent_company":             comp_res.get("parent_company",""),
        "domestic_ultimate":          comp_res.get("domestic_ultimate",""),
        "global_ultimate":            comp_res.get("global_ultimate",""),
        "is_headquarters":            comp_res.get("is_headquarters",""),
        "email_domain":               comp_res.get("email_domain",""),
        "naics_code":                 comp_res.get("naics_code",""),
        "naics_description":          comp_res.get("naics_description",""),
        "sic_code":                   comp_res.get("sic_code",""),
        "sic_description":            comp_res.get("sic_description",""),
        "employee_count_range":       comp_res.get("employee_count_range",""),
        "revenue_range":              comp_res.get("revenue_range",""),
        "year_established":           comp_res.get("year_established",""),
        "industry":                   comp_res.get("industry",""),
        "enrich_source":              comp_res.get("enrich_source",""),
        # sources (URLs the AI grounded against — from web_search_preview citations)
        "sources":                    _format_sources(combined),
        "sources_verify":             _format_sources(verify_srcs),
        "sources_enrich":             _format_sources(enrich_srcs),
        "sources_hq":                 _format_sources(hq_srcs),
        # headquarters
        "headquarters_address":       comp_res.get("headquarters_address",""),
        "headquarters_source":        comp_res.get("headquarters_source",""),
        # meta
        "worker_id":              worker_id,
        "pipeline_latency_sec":   elapsed,
    }

    # ── Worker-visible log line (shows thread assignment) ────────────────────
    print(f"  [{worker_id}] MDM={rec['MDM_KEY']:<12} | "
          f"addr={addr_res.get('addr_final_status','?'):<22} | "
          f"route={decision:<11} | "
          f"company={comp_res['match_status']:<30} | "
          f"{elapsed}s")

    return out

# ═══════════════════════════════════════════════════════════════════════════════
# ██  WRITER THREAD
# ═══════════════════════════════════════════════════════════════════════════════
def _writer(output_csv, open_mode, status_counts, routing_counts,
            processed_box, batch_box, in_batch_box, total, pbar):
    fh = open(output_csv, mode=open_mode, newline="", encoding="utf-8-sig")
    wr = csv.DictWriter(fh, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
    if open_mode == "w":
        wr.writeheader(); fh.flush()

    while True:
        try: item = _wq.get(timeout=2)
        except Empty: continue
        if item is _SENTINEL: break
        wr.writerow(item); fh.flush()

        with _cl:
            s = item.get("match_status","")
            status_counts[s] = status_counts.get(s,0) + 1
            r = item.get("routing_decision","")
            routing_counts[r] = routing_counts.get(r,0) + 1
            processed_box[0] += 1
            in_batch_box[0]  += 1
            if in_batch_box[0] >= BATCH_SIZE:
                batch_box[0] += 1
                _banner(batch_box[0], status_counts, routing_counts, processed_box[0], total)
                in_batch_box[0] = 0
        pbar.update(1)
    fh.close()

def _banner(bn, sc, rc, done, total):
    found  = sum(v for k,v in sc.items() if "COMPANY_FOUND" in k and "DIFF" not in k and "CACHED" not in k)
    diff   = sum(v for k,v in sc.items() if "DIFF_ADDR" in k)
    nf     = sum(v for k,v in sc.items() if "NOT_FOUND" in k)
    skip   = sum(v for k,v in sc.items() if "SKIPPED" in k)
    cached = sum(v for k,v in sc.items() if "CACHED" in k)
    full_m = rc.get("FULL_MATCH", 0)
    name_f = rc.get("NAME_FIRST", 0)
    skip_r = rc.get("SKIP", 0)
    pct    = 100*done/total if total else 0
    print("\n" + "▓"*70)
    print(f"  BATCH {bn} | {done}/{total} ({pct:.1f}%)  cache_hits={cached}")
    print("▓"*70)
    print(f"  ✅ COMPANY_FOUND           : {found}")
    print(f"  🔄 COMPANY_FOUND_DIFF_ADDR : {diff}")
    print(f"  ❌ COMPANY_NOT_FOUND       : {nf}")
    print(f"  ⏭  SKIPPED                : {skip}")
    print(f"  ⚡ CACHED                  : {cached}")
    print(f"  ─── ROUTING ─────────────────")
    print(f"  🛣  FULL_MATCH              : {full_m}")
    print(f"  🎯 NAME_FIRST              : {name_f}")
    print(f"  🚫 SKIP                    : {skip_r}")
    print("▓"*70 + "\n")

# ═══════════════════════════════════════════════════════════════════════════════
# ██  SUMMARY REPORT
# ═══════════════════════════════════════════════════════════════════════════════
def _pct(n, total):
    return f"{100*n/total:.1f}%" if total else "0.0%"

def _section(title: str) -> str:
    return f"\n{'═'*70}\n  {title}\n{'═'*70}\n"

def generate_summary_report(
    output_csv: str,
    summary_path: Optional[str] = None,
    review_queue_path: Optional[str] = None,
    wall_time_sec: float = 0.0,
) -> Optional[str]:
    """Reads the output CSV and produces a stage-by-stage summary plus a manual
    review queue CSV. Called at end of run(). Returns the body string for downstream
    use (e.g. narrative summarization), or None if the output couldn't be read."""
    if not os.path.isfile(output_csv):
        logging.error(f"Summary: output file {output_csv} not found")
        return None

    df = pd.read_csv(output_csv, encoding="utf-8-sig", low_memory=False)
    total = len(df)
    if total == 0:
        logging.warning("Summary: output CSV is empty")
        return None

    lines = []
    lines.append(_section("MDM PIPELINE — SUMMARY REPORT"))
    lines.append(f"  Output file       : {output_csv}")
    lines.append(f"  Records processed : {total}")
    lines.append(f"  Wall time         : {wall_time_sec:.1f}s ({total/wall_time_sec:.1f} rec/s)" if wall_time_sec > 0 else "")

    # ── Pre-processing ────────────────────────────────────────────────────────
    lines.append(_section("STAGE 1 — Pre-processing"))
    skip_count = int(df["skip_api"].astype(str).str.lower().eq("true").sum()) if "skip_api" in df else 0
    moj_count  = int(df["mojibake_rescued"].astype(str).str.lower().eq("true").sum()) if "mojibake_rescued" in df else 0
    trans_count= int(df["street_translated"].astype(str).str.lower().eq("true").sum()) if "street_translated" in df else 0
    lines.append(f"  Mojibake rescued  : {moj_count} ({_pct(moj_count,total)})")
    lines.append(f"  Street translated : {trans_count} ({_pct(trans_count,total)})")
    lines.append(f"  Skipped (no addr) : {skip_count} ({_pct(skip_count,total)})")
    if "pre_flags" in df:
        flag_counts = {}
        for f in df["pre_flags"].fillna("").astype(str):
            for tag in f.split("|"):
                tag = tag.strip()
                if tag and tag != "OK":
                    flag_counts[tag] = flag_counts.get(tag, 0) + 1
        if flag_counts:
            lines.append(f"  Flag distribution :")
            for tag, n in sorted(flag_counts.items(), key=lambda x:-x[1]):
                lines.append(f"    {tag:<32} {n:>5}  ({_pct(n,total)})")

    # ── Address validation ────────────────────────────────────────────────────
    lines.append(_section("STAGE 2 — Address validation (Azure Maps)"))
    if "addr_final_status" in df:
        status_counts = df["addr_final_status"].fillna("").value_counts()
        for s, n in status_counts.items():
            if s:
                lines.append(f"  {str(s):<28} {n:>5}  ({_pct(n,total)})")
    if "val_score" in df:
        scores = pd.to_numeric(df["val_score"], errors="coerce").dropna()
        high   = int((scores >= 0.85).sum())
        mid    = int(((scores >= 0.6) & (scores < 0.85)).sum())
        low    = int((scores < 0.6).sum())
        lines.append(f"  Score >= 0.85  (gate pass) : {high} ({_pct(high,total)})")
        lines.append(f"  Score 0.6-0.85 (medium)    : {mid} ({_pct(mid,total)})")
        lines.append(f"  Score <  0.6   (low/zero)  : {low} ({_pct(low,total)})")
    if "addr_correction_strategy" in df:
        strat = df["addr_correction_strategy"].fillna("").value_counts()
        lines.append(f"  Correction strategy used :")
        for s, n in strat.items():
            if s:
                lines.append(f"    {str(s):<28} {n:>5}  ({_pct(n,total)})")

    # ── Routing ───────────────────────────────────────────────────────────────
    lines.append(_section("STAGE 3 — Routing decision"))
    if "routing_decision" in df:
        for d, n in df["routing_decision"].fillna("").value_counts().items():
            if d:
                lines.append(f"  {str(d):<14} {n:>5}  ({_pct(n,total)})")

    # ── Company match ─────────────────────────────────────────────────────────
    lines.append(_section("STAGE 4 — Company match"))
    if "match_status" in df:
        for s, n in df["match_status"].fillna("").value_counts().items():
            if s:
                lines.append(f"  {str(s):<35} {n:>5}  ({_pct(n,total)})")
    if "tier_used" in df:
        lines.append(f"\n  Tier breakdown :")
        for t, n in df["tier_used"].fillna("").value_counts().items():
            if t:
                lines.append(f"    {str(t):<28} {n:>5}  ({_pct(n,total)})")

    # ── Enrichment coverage (only verified records) ───────────────────────────
    lines.append(_section("STAGE 5 — Enrichment coverage (verified records only)"))
    verified_mask = df.get("company_exists","").astype(str).str.lower().eq("true")
    n_verified = int(verified_mask.sum())
    lines.append(f"  Verified records  : {n_verified} ({_pct(n_verified,total)})")
    if n_verified > 0:
        verified = df[verified_mask]
        for f in ENRICHMENT_FIELDS:
            if f in verified.columns:
                filled = int(verified[f].fillna("").astype(str).str.strip().ne("").sum())
                lines.append(f"  {f:<24} {filled:>5} / {n_verified:<5}  ({_pct(filled,n_verified)})")

    # ── Headquarters resolution coverage ──────────────────────────────────────
    if "headquarters_address" in df.columns:
        lines.append(_section("STAGE 5b — Headquarters resolution"))
        if n_verified > 0:
            verified = df[verified_mask]
            hq_filled = int(verified["headquarters_address"].fillna("").astype(str).str.strip().ne("").sum())
            lines.append(f"  HQ address resolved : {hq_filled} / {n_verified}  ({_pct(hq_filled,n_verified)})")
            # Source breakdown
            src_series = verified["headquarters_source"].fillna("").astype(str)
            # Strip the trailing ":target_name" / "via web_search" so we get clean buckets
            src_buckets = src_series.str.split(":", n=1).str[0].str.split(" via ").str[0]
            src_counts = src_buckets[src_buckets != ""].value_counts()
            for src, n in src_counts.items():
                lines.append(f"    {str(src):<28} {n:>5}  ({_pct(n,n_verified)})")
        lines.append(f"  Unique HQ targets cached : {len(_hq_cache)}")

    # ── Manual review queue ───────────────────────────────────────────────────
    review_mask = (
        df.get("ai_confidence","").astype(str).str.upper().eq("LOW")
        | df.get("address_match","").astype(str).str.upper().eq("UNKNOWN")
        | df.get("match_status","").astype(str).str.contains("UNVERIFIED|NOT_FOUND|MANUAL_REVIEW", na=False, regex=True)
        | df.get("addr_final_status","").astype(str).str.contains("MANUAL_REVIEW|API_ERROR|INVALID", na=False, regex=True)
    )
    review_df = df[review_mask]
    if review_queue_path and len(review_df) > 0:
        try:
            review_df.to_csv(review_queue_path, index=False, encoding="utf-8-sig")
        except Exception as e:
            logging.error(f"Failed to write review queue: {e}")
    lines.append(_section("STAGE 6 — Manual review queue"))
    lines.append(f"  Records flagged for manual review : {len(review_df)} ({_pct(len(review_df),total)})")
    if review_queue_path:
        lines.append(f"  Written to : {review_queue_path}")

    # ── AI calls + cost estimate ──────────────────────────────────────────────
    lines.append(_section("AI calls (actual)"))
    v_api    = _ai_calls.get("verify_api", 0)
    v_fail   = _ai_calls.get("verify_failed", 0)
    e_api    = _ai_calls.get("enrich_api", 0)
    e_cached = _ai_calls.get("enrich_cached", 0)
    h_api    = _ai_calls.get("hq_api", 0)
    h_cached = _ai_calls.get("hq_cached", 0)
    n_narr   = _ai_calls.get("narrative", 0)
    total_billable = v_api + e_api + h_api + n_narr
    lines.append(f"  Tier 3 verify (web search) : {v_api} API" + (f" ({v_fail} failed)" if v_fail else ""))
    lines.append(f"  Enrichment (web search)    : {e_api} API  +  {e_cached} cache hits")
    lines.append(f"  HQ lookup (web search)     : {h_api} API  +  {h_cached} cache hits")
    lines.append(f"  Narrative summary          : {n_narr}")
    lines.append(f"  Total billable AI calls    : {total_billable}")
    lines.append(f"  Cache savings (avoided)    : {e_cached + h_cached} calls")
    lines.append(f"  Unique enrichment keys     : {len(_enrich_cache)}")
    lines.append(f"  Unique HQ keys             : {len(_hq_cache)}")
    lines.append(f"  Unique address-match keys  : {len(_cache)}")
    lines.append(f"  Estimated OpenAI spend     : ${_ai_calls_estimated_spend():.2f}  "
                 f"(verify+enrich+hq @ ~$0.20, narrative @ ~$0.01)")

    # ── Throughput ────────────────────────────────────────────────────────────
    if "pipeline_latency_sec" in df:
        lat = pd.to_numeric(df["pipeline_latency_sec"], errors="coerce").dropna()
        if len(lat) > 0:
            lines.append(_section("Per-record latency"))
            lines.append(f"  p50 : {lat.median():.2f}s")
            lines.append(f"  p95 : {lat.quantile(0.95):.2f}s")
            lines.append(f"  max : {lat.max():.2f}s")

    body = "\n".join(l for l in lines if l is not None)

    print(body)
    if summary_path:
        try:
            with open(summary_path, "w", encoding="utf-8") as fh:
                fh.write(body + "\n")
            print(f"\n  📄 Summary saved → {summary_path}")
        except Exception as e:
            logging.error(f"Failed to write summary: {e}")
    return body

# ═══════════════════════════════════════════════════════════════════════════════
# ██  NARRATIVE SUMMARY  (one OpenAI call to turn stats into prose)
# ═══════════════════════════════════════════════════════════════════════════════
NARRATIVE_SYSTEM = """You are a senior data engineering lead writing the executive summary
of a Master Data Management (MDM) pipeline run. The pipeline takes raw, malformed
customer records and uses Azure Maps + OpenAI web search to validate addresses,
verify company presence, and enrich with firmographics (NAICS, SIC, hierarchy, etc.).

Write 2-3 short paragraphs (max 250 words total) covering:
1. Headline result — how many records were verified vs. failed, and what % is workable
2. Top 2-3 failure modes and their likely root causes (preprocessing flags, address quality, name quality)
3. Enrichment coverage — which firmographic fields populated well vs. poorly, and why
4. One concrete recommendation to improve the next run (e.g. raise threshold, add a flag, run subset manually)

Style: direct, factual, no fluff. No marketing language. No bullet points unless absolutely needed.
Cite specific numbers from the stats. If a metric is missing, just don't mention it."""

def generate_narrative_summary(
    summary_body: str,
    narrative_path: Optional[str] = None,
) -> Optional[str]:
    """One OpenAI call that turns the stage-by-stage summary into a 2-3 paragraph
    executive narrative. No web search — pure reasoning over the stats."""
    if not summary_body:
        return None
    try:
        rsp = oai_client.responses.create(
            model=MODEL_NAME,
            input=[
                {"role":"system","content":NARRATIVE_SYSTEM},
                {"role":"user","content":f"Here are the stage-by-stage stats from the run:\n\n{summary_body}\n\nWrite the executive summary."},
            ],
        )
        _bump_ai("narrative")
        narrative = rsp.output_text.strip()
    except Exception as e:
        logging.error(f"Narrative summary failed: {e}")
        return None

    print("\n" + "═"*70)
    print("  📝 EXECUTIVE NARRATIVE (LLM-generated)")
    print("═"*70 + "\n")
    print(narrative)
    if narrative_path:
        try:
            with open(narrative_path, "w", encoding="utf-8") as fh:
                fh.write(narrative + "\n")
            print(f"\n  📄 Narrative saved → {narrative_path}")
        except Exception as e:
            logging.error(f"Failed to write narrative: {e}")
    return narrative

# ═══════════════════════════════════════════════════════════════════════════════
# ██  EXCEL REPORT  (single multi-sheet xlsx for human review)
# ═══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(
    output_csv: str,
    xlsx_path: str,
    wall_time_sec: float = 0.0,
    summary_body: Optional[str] = None,
    narrative_text: Optional[str] = None,
) -> Optional[str]:
    """One xlsx workbook with all run artifacts as sheets:
      Overview            — headline metrics, AI call counts, estimated spend
      Narrative           — LLM executive summary (if supplied)
      Summary             — full stage-by-stage stats text (if supplied)
      By Stage            — per-stage value-counts (validation / routing / match / tier / flags)
      Records             — every record, key columns, conditional formatting + auto-sized
      Manual Review       — only flagged records (LOW conf / NOT_FOUND / DOWNGRADED / errors)
      Enrichment Coverage — fill rate per firmographic field on verified records
    Column widths are auto-sized from actual content (capped at 60 chars); long-text
    columns wrap; header row frozen; auto-filter enabled on data sheets.
    Returns the xlsx path on success, None otherwise.
    """
    if not HAVE_OPENPYXL:
        logging.error("openpyxl not installed — xlsx report skipped. pip install openpyxl")
        return None
    if not os.path.isfile(output_csv):
        logging.error(f"xlsx report: output {output_csv} not found")
        return None

    df = pd.read_csv(output_csv, encoding="utf-8-sig", low_memory=False)
    total = len(df)
    if total == 0:
        return None

    for c in ("val_score", "confidence_score", "pipeline_latency_sec"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    wb = Workbook(); wb.remove(wb.active)

    # ── Style palette ──────────────────────────────────────────────────────
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    H1_FONT  = Font(bold=True, size=14, color="1F4E79")
    H2_FONT  = Font(bold=True, size=12, color="1F4E79")
    THIN     = Side(style="thin", color="BFBFBF")
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER   = Alignment(horizontal="center", vertical="center")
    LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    WRAP     = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    GREEN    = PatternFill("solid", fgColor="C6EFCE")
    YELLOW   = PatternFill("solid", fgColor="FFEB9C")
    RED      = PatternFill("solid", fgColor="FFC7CE")
    GRAY     = PatternFill("solid", fgColor="D9D9D9")

    # Long-text columns: wrap text, generous width; everything else auto-sized.
    WRAP_COLS = {"FULL_ADDRESS_RAW", "val_returned_freeform", "confidence_reason",
                 "ai_evidence", "all_known_locations", "best_address",
                 "headquarters_address", "headquarters_source",
                 "sources", "sources_verify", "sources_enrich", "sources_hq",
                 "naics_description", "industry"}
    MIN_WIDTH = 8
    MAX_WIDTH = 60
    WRAP_WIDTH = 50

    # Excel forbids ASCII control chars (\x00-\x08, \x0b-\x0c, \x0e-\x1f).
    # OpenAI responses can include them when the model emits raw bytes; openpyxl
    # raises IllegalCharacterError on write. Strip them defensively.
    _XLSX_BAD_CHARS = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
    XLSX_CELL_LIMIT = 32760  # Excel hard limit is 32767 chars per cell; leave headroom

    def _safe(v):
        if isinstance(v, str):
            v = _XLSX_BAD_CHARS.sub("", v)
            if len(v) > XLSX_CELL_LIMIT:
                v = v[:XLSX_CELL_LIMIT] + "…[truncated]"
        return v

    def _pct(n, denom):
        return f"{100*n/denom:.1f}%" if denom else "0.0%"

    def _autosize(ws, headers, df_in, hdr_row=1):
        """Compute column widths from actual content (capped). Wrap-text columns
        get a fixed wrap width so they look uniform."""
        for j, h in enumerate(headers, 1):
            if h in WRAP_COLS:
                w = WRAP_WIDTH
            else:
                # max(header, longest cell) + small padding
                col_vals = df_in[h].astype(str).fillna("") if h in df_in.columns else pd.Series([""])
                longest = max(len(h), int(col_vals.str.len().max() or 0))
                w = min(max(longest + 2, MIN_WIDTH), MAX_WIDTH)
            ws.column_dimensions[get_column_letter(j)].width = w

    def _write_kv_table(ws, start_row, header, rows, col_widths=None):
        """Generic header + rows table for the Overview / By-Stage / Coverage sheets."""
        for j, h in enumerate(header, 1):
            c = ws.cell(row=start_row, column=j, value=_safe(h))
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
        r = start_row + 1
        for row in rows:
            for j, v in enumerate(row, 1):
                c = ws.cell(row=r, column=j, value=_safe(v))
                c.border = BORDER; c.alignment = LEFT
            r += 1
        if col_widths:
            for j, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(j)].width = w
        return r + 1

    def _write_records(ws, df_in, headers, hdr_row=1):
        """Write a records sheet: bold blue header, wrap on long-text cols,
        frozen pane below header, auto-filter, and auto-sized columns."""
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=hdr_row, column=j, value=_safe(h))
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER
        for i, (_, row) in enumerate(df_in.iterrows(), start=hdr_row + 1):
            for j, h in enumerate(headers, 1):
                v = row.get(h, "")
                if pd.isna(v): v = None
                cell = ws.cell(row=i, column=j, value=_safe(v))
                if h in WRAP_COLS:
                    cell.alignment = WRAP
                else:
                    cell.alignment = LEFT
        _autosize(ws, headers, df_in, hdr_row)
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1).coordinate
        last = hdr_row + len(df_in)
        if last > hdr_row:
            ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(headers))}{last}"

    def _apply_record_formatting(ws, n_rows, headers, hdr_row=1):
        """Color match_status (green/yellow/red/gray) + color-scale confidence_score."""
        if "match_status" in headers:
            cl = get_column_letter(headers.index("match_status") + 1)
            rng = f"{cl}{hdr_row+1}:{cl}{hdr_row+n_rows}"
            anchor = f"{cl}{hdr_row+1}"
            for formula, fill in [
                (f'ISNUMBER(SEARCH("DIFF_ADDR",{anchor}))',     YELLOW),
                (f'ISNUMBER(SEARCH("NOT_FOUND",{anchor}))',     RED),
                (f'ISNUMBER(SEARCH("SKIPPED",{anchor}))',       GRAY),
                (f'ISNUMBER(SEARCH("COMPANY_FOUND",{anchor}))', GREEN),
            ]:
                ws.conditional_formatting.add(rng,
                    FormulaRule(formula=[formula], fill=fill, stopIfTrue=True))
        if "confidence_score" in headers:
            cl = get_column_letter(headers.index("confidence_score") + 1)
            rng = f"{cl}{hdr_row+1}:{cl}{hdr_row+n_rows}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0,   start_color="F8696B",
                mid_type="num",   mid_value=0.5,   mid_color="FFEB84",
                end_type="num",   end_value=1.0,   end_color="63BE7B",
            ))

    # ── Sheet 1: Overview ──────────────────────────────────────────────────
    ws = wb.create_sheet("Overview")
    ws["A1"] = "MDM Pipeline — Run Overview"; ws["A1"].font = H1_FONT

    verified_mask = df.get("company_exists","").astype(str).str.lower().eq("true")
    n_verified = int(verified_mask.sum())
    ms = df.get("match_status", pd.Series(dtype=str)).fillna("").astype(str)
    n_diff   = int(ms.str.contains("DIFF_ADDR", na=False).sum())
    n_nf     = int(ms.str.contains("NOT_FOUND", na=False).sum())
    n_skip   = int(ms.str.contains("SKIPPED", na=False).sum())
    n_cached = int(ms.str.contains("CACHED", na=False).sum())

    addr_status = df.get("addr_final_status", pd.Series(dtype=str)).fillna("").astype(str)
    n_addr_pass = int(addr_status.isin(["VALID","VALID_LOW_PRECISION"]).sum())
    n_addr_fail = int(addr_status.isin(["API_ERROR","INVALID","NO_RESULT","MANUAL_REVIEW_NEEDED"]).sum())

    hq_filled = 0
    if "headquarters_address" in df.columns:
        hq_filled = int(df["headquarters_address"].fillna("").astype(str).str.strip().ne("").sum())

    # Pipeline-downgrade count (verdict-rigor guard)
    n_downgraded = 0
    if "ai_evidence" in df.columns:
        n_downgraded = int(df["ai_evidence"].fillna("").astype(str)
                           .str.contains(r"PIPELINE DOWNGRADE", regex=True).sum())

    overview_rows = [
        ["Records processed",  total],
        ["Wall time (sec)",    round(wall_time_sec, 1) if wall_time_sec else ""],
        ["Throughput (rec/s)", round(total/wall_time_sec, 2) if wall_time_sec else ""],
        ["", ""],
        ["Verified (company_exists=True)", f"{n_verified}  ({_pct(n_verified, total)})"],
        ["Found at different address",     f"{n_diff}  ({_pct(n_diff, total)})"],
        ["Not found",                      f"{n_nf}  ({_pct(n_nf, total)})"],
        ["Skipped",                        f"{n_skip}  ({_pct(n_skip, total)})"],
        ["Cache hits",                     f"{n_cached}  ({_pct(n_cached, total)})"],
        ["", ""],
        ["Address validation pass",        f"{n_addr_pass}  ({_pct(n_addr_pass, total)})"],
        ["Address errors / no-result",     f"{n_addr_fail}  ({_pct(n_addr_fail, total)})"],
        ["", ""],
        ["Headquarters resolved",          f"{hq_filled}  ({_pct(hq_filled, n_verified)} of verified)"],
        ["Pipeline-downgraded verdicts",   n_downgraded],
        ["", ""],
        ["── AI calls (actual) ──", ""],
        ["Tier 3 verify (API)",        _ai_calls.get("verify_api", 0)],
        ["Tier 3 verify (failed)",     _ai_calls.get("verify_failed", 0)],
        ["Enrichment (API)",           _ai_calls.get("enrich_api", 0)],
        ["Enrichment (cache hits)",    _ai_calls.get("enrich_cached", 0)],
        ["HQ lookup (API)",            _ai_calls.get("hq_api", 0)],
        ["HQ lookup (cache hits)",     _ai_calls.get("hq_cached", 0)],
        ["Narrative summary",          _ai_calls.get("narrative", 0)],
        ["Total billable AI calls",    _ai_calls_total_billable()],
        ["Cache savings (avoided)",    _ai_calls.get("enrich_cached", 0) + _ai_calls.get("hq_cached", 0)],
        ["Estimated OpenAI spend",     f"${_ai_calls_estimated_spend():.2f}"],
    ]
    _write_kv_table(ws, 3, ["Metric", "Value"], overview_rows, col_widths=[42, 32])
    ws.freeze_panes = "A4"

    # ── Sheet 2: Narrative (LLM exec summary) ──────────────────────────────
    if narrative_text:
        ws = wb.create_sheet("Narrative")
        ws["A1"] = "Executive Narrative (LLM-generated)"
        ws["A1"].font = H1_FONT
        # write paragraphs as separate rows so wrap renders cleanly
        ws.column_dimensions["A"].width = 110
        r = 3
        for para in narrative_text.strip().split("\n\n"):
            cell = ws.cell(row=r, column=1, value=_safe(para.strip()))
            cell.alignment = WRAP
            ws.row_dimensions[r].height = max(40, min(20 * (len(para) // 90 + 1), 300))
            r += 2
        ws.freeze_panes = "A2"

    # ── Sheet 3: Summary (stage-by-stage stats text) ───────────────────────
    if summary_body:
        ws = wb.create_sheet("Summary")
        ws["A1"] = "Stage-by-stage Summary"
        ws["A1"].font = H1_FONT
        ws.column_dimensions["A"].width = 100
        # write line by line so the user can scroll like a text file
        for i, line in enumerate(summary_body.splitlines(), start=3):
            cell = ws.cell(row=i, column=1, value=_safe(line))
            cell.alignment = LEFT
            cell.font = Font(name="Consolas", size=10)
        ws.freeze_panes = "A2"

    # ── Sheet 4: By Stage ──────────────────────────────────────────────────
    ws = wb.create_sheet("By Stage")
    ws["A1"] = "MDM Pipeline — Stage Breakdowns"; ws["A1"].font = H1_FONT
    r = 3

    def _stage_section(start_r, title, series, denom):
        ws.cell(row=start_r, column=1, value=title).font = H2_FONT
        rows = [[str(k), int(n), _pct(int(n), denom)]
                for k, n in series.fillna("").value_counts().items() if k]
        return _write_kv_table(ws, start_r + 1, ["Bucket", "Count", "%"], rows,
                               col_widths=[44, 12, 12])

    for title, col in [
        ("Address validation status",   "addr_final_status"),
        ("Address correction strategy", "addr_correction_strategy"),
        ("Routing decision",            "routing_decision"),
        ("Match status",                "match_status"),
        ("Tier used",                   "tier_used"),
        ("AI confidence",               "ai_confidence"),
    ]:
        if col in df.columns:
            r = _stage_section(r, title, df[col].astype(str), total)

    if "pre_flags" in df.columns:
        ws.cell(row=r, column=1, value="Pre-processing flags").font = H2_FONT
        flag_counts: Dict[str, int] = {}
        for f in df["pre_flags"].fillna("").astype(str):
            for tag in f.split("|"):
                tag = tag.strip()
                if tag and tag != "OK":
                    flag_counts[tag] = flag_counts.get(tag, 0) + 1
        flag_rows = sorted(([k, v, _pct(v, total)] for k, v in flag_counts.items()),
                           key=lambda x: -x[1])
        r = _write_kv_table(ws, r + 1, ["Flag", "Count", "%"], flag_rows,
                            col_widths=[40, 12, 12])

    # ── Sheet 3: Records ───────────────────────────────────────────────────
    KEY_COLS = [
        "MDM_KEY","SOURCE_NAME","country","FULL_ADDRESS_RAW",
        "addr_final_status","val_score","val_returned_freeform",
        "routing_decision","match_status","tier_used",
        "canonical_name","company_exists","address_match",
        "ai_confidence","confidence_score","confidence_reason","ai_evidence",
        "all_known_locations","best_address","mdm_address_occupant",
        "naics_code","naics_description","industry",
        "headquarters_address","headquarters_source",
        "sources","sources_verify","sources_enrich","sources_hq",
        "pipeline_latency_sec",
    ]
    headers = [c for c in KEY_COLS if c in df.columns]

    ws = wb.create_sheet("Records")
    _write_records(ws, df, headers, hdr_row=1)
    _apply_record_formatting(ws, len(df), headers, hdr_row=1)

    # ── Sheet 4: Manual Review ─────────────────────────────────────────────
    review_mask = (
        df.get("ai_confidence", pd.Series(dtype=str)).astype(str).str.upper().eq("LOW")
        | df.get("address_match", pd.Series(dtype=str)).astype(str).str.upper().eq("UNKNOWN")
        | df.get("match_status", pd.Series(dtype=str)).astype(str).str.contains(
            "UNVERIFIED|NOT_FOUND|MANUAL_REVIEW", na=False, regex=True)
        | df.get("addr_final_status", pd.Series(dtype=str)).astype(str).str.contains(
            "MANUAL_REVIEW|API_ERROR|INVALID", na=False, regex=True)
        | df.get("ai_evidence", pd.Series(dtype=str)).astype(str).str.contains(
            "PIPELINE DOWNGRADE", na=False, regex=True)
    )
    rdf = df[review_mask].copy()
    ws = wb.create_sheet("Manual Review")
    ws["A1"] = f"Records flagged for manual review ({len(rdf)} of {total})"
    ws["A1"].font = H1_FONT
    if len(rdf) > 0:
        _write_records(ws, rdf, headers, hdr_row=3)
        _apply_record_formatting(ws, len(rdf), headers, hdr_row=3)
    else:
        ws["A3"] = "No records flagged — all verified or skipped cleanly."

    # ── Sheet 5: Enrichment Coverage ───────────────────────────────────────
    ws = wb.create_sheet("Enrichment Coverage")
    ws["A1"] = f"Enrichment fill rate (verified records: {n_verified})"
    ws["A1"].font = H1_FONT
    if n_verified > 0:
        v_df = df[verified_mask]
        rows = []
        for f in ENRICHMENT_FIELDS:
            if f in v_df.columns:
                filled = int(v_df[f].fillna("").astype(str).str.strip().ne("").sum())
                rows.append([f, filled, n_verified, _pct(filled, n_verified)])
        if "headquarters_address" in v_df.columns:
            filled = int(v_df["headquarters_address"].fillna("").astype(str).str.strip().ne("").sum())
            rows.append(["headquarters_address", filled, n_verified, _pct(filled, n_verified)])
        _write_kv_table(ws, 3, ["Field", "Filled", "Total verified", "%"], rows,
                        col_widths=[30, 12, 18, 12])
    ws.freeze_panes = "A4"

    try:
        wb.save(xlsx_path)
        print(f"\n  Excel report saved → {xlsx_path}")
        return xlsx_path
    except Exception as e:
        logging.error(f"Failed to save xlsx {xlsx_path}: {e}")
        return None

# ═══════════════════════════════════════════════════════════════════════════════
# ██  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def run(input_csv: str, output_csv: str, test_mode: bool = False):
    run_t0 = time.time()
    df    = pd.read_csv(input_csv, encoding="utf-8-sig")
    total = len(df)
    n_workers = MAX_WORKERS

    if test_mode:
        df        = df.head(TEST_MODE_ROWS)
        total     = len(df)
        n_workers = min(MAX_WORKERS, total)   # don't spin more threads than records
        print(f"\n{'='*70}")
        print(f"  🧪 TEST MODE — {total} records × {n_workers} workers")
        print(f"  Each worker will log its thread name beside every record.")
        print(f"  You should see up to {n_workers} unique worker IDs below.")
        print(f"{'='*70}\n")
    else:
        print(f"\n{'='*70}")
        print(f"  🚀 FULL RUN — {total} records × {n_workers} workers")
        print(f"  rapidfuzz : {'✅' if HAVE_RAPIDFUZZ else '⚠️  not installed'}")
        print(f"  langdetect: {'✅' if HAVE_LANGDETECT else '⚠️  not installed'}")
        print(f"  translator: {'✅' if HAVE_TRANSLATOR else '⚠️  not installed'}")
        print(f"{'='*70}\n")

    # Resume
    already_done: set = set()
    file_exists = os.path.isfile(output_csv) and not test_mode
    if file_exists:
        try:
            done_df      = pd.read_csv(output_csv, encoding="utf-8-sig", usecols=["MDM_KEY"])
            already_done = set(str(v) for v in done_df["MDM_KEY"].dropna())
            logging.info(f"Resume — {len(already_done)} already done, skipping.")
        except: file_exists = False

    open_mode = "a" if file_exists and already_done else "w"
    rows = [row for _, row in df.iterrows()
            if str(row.get("MDM_KEY","")) not in already_done]
    logging.info(f"Records to process: {len(rows)}")

    sc = {}
    rc = {}
    pb = [len(already_done)]
    bb = [0]; ib = [0]
    pbar = tqdm(total=total, initial=len(already_done), desc="Pipeline", unit="rec")

    wt = threading.Thread(
        target=_writer,
        args=(output_csv, open_mode, sc, rc, pb, bb, ib, total, pbar),
        daemon=True, name="csv-writer"
    )
    wt.start()

    with ThreadPoolExecutor(max_workers=n_workers, thread_name_prefix="mdm") as ex:
        futs = {ex.submit(process_record, row): row for row in rows}
        for fut in as_completed(futs):
            try:    _wq.put(fut.result())
            except Exception as e:
                logging.error(f"Worker crash: {e}")
                pbar.update(1)

    _wq.put(_SENTINEL)
    wt.join()
    pbar.close()

    if ib[0] > 0:
        bb[0] += 1
        _banner(bb[0], sc, rc, pb[0], total)

    print(f"\n✅  Output → {output_csv}")
    print(f"   Cache size       : {len(_cache)} addr-keyed pairs")
    print(f"   Name-cache size  : {len(_name_cache)} (name, country) pairs")
    print(f"   Enrich-cache size: {len(_enrich_cache)} (canonical_name, country) pairs")
    print(f"   HQ-cache size    : {len(_hq_cache)} (parent/ultimate, country) pairs")
    print(f"   AI calls         : verify={_ai_calls.get('verify_api',0)} "
          f"enrich={_ai_calls.get('enrich_api',0)}+{_ai_calls.get('enrich_cached',0)}c "
          f"hq={_ai_calls.get('hq_api',0)}+{_ai_calls.get('hq_cached',0)}c "
          f"narrative={_ai_calls.get('narrative',0)}  →  ${_ai_calls_estimated_spend():.2f}")
    for k, v in sorted(sc.items(), key=lambda x:-x[1]):
        print(f"   {k:<45} {v:>5}  ({100*v/pb[0]:.1f}%)" if pb[0] else f"   {k}: {v}")

    # ── Single consolidated xlsx report ───────────────────────────────────────
    # All run artifacts (summary text, LLM narrative, review queue, stats,
    # records) go into one workbook as separate sheets — no auxiliary files.
    summary_base = output_csv.rsplit(".", 1)[0]
    xlsx_path    = f"{summary_base}.xlsx"
    summary_body = generate_summary_report(
        output_csv=output_csv,
        wall_time_sec=time.time() - run_t0,
    )
    narrative_text = None
    if summary_body:
        narrative_text = generate_narrative_summary(summary_body)
    generate_excel_report(
        output_csv=output_csv,
        xlsx_path=xlsx_path,
        wall_time_sec=time.time() - run_t0,
        summary_body=summary_body,
        narrative_text=narrative_text,
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="MDM Full Pipeline v9")
    parser.add_argument("--test",   action="store_true", help=f"Run first {TEST_MODE_ROWS} records only")
    parser.add_argument("--input",  default=INPUT_CSV,   help="Input CSV path")
    parser.add_argument("--output", default=OUTPUT_CSV,  help="Output CSV path")
    args = parser.parse_args()

    run(input_csv=args.input, output_csv=args.output, test_mode=args.test)
